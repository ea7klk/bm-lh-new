from __future__ import annotations

import asyncio
from contextlib import suppress
from datetime import datetime, timedelta, timezone
import csv
import html
import hmac
import io
import json
import re
from time import perf_counter
from urllib.parse import parse_qs, quote, urlencode
from pathlib import Path
from typing import Any

import psycopg
from fastapi import FastAPI, Query, Request, WebSocket, WebSocketDisconnect
from fastapi.encoders import jsonable_encoder
from fastapi.responses import HTMLResponse, JSONResponse, RedirectResponse
from psycopg.errors import UniqueViolation
from starlette.responses import Response

from .auth import (
    hash_password,
    is_bcrypt_hash,
    issue_admin_token,
    new_email_verification_token,
    new_password_reset_token,
    new_session_token,
    session_token_hash,
    verify_admin_token,
    verify_password,
)
from .config import settings
from .consent import cookie_consent_markup, cookie_consent_script
from .email import (
    EmailDeliveryError,
    send_password_reset_email,
    send_verification_email,
)
from .i18n import (
    LANGUAGE_COOKIE,
    LANGUAGE_COOKIE_MAX_AGE,
    LANGUAGE_INFO,
    SUPPORTED_LOCALES,
    catalog,
    normalize_locale,
    translate,
)
from .matomo import matomo_configured, matomo_script
from .storage import QSO_NOTIFY_CHANNEL, PostgresStore


UTC = timezone.utc
TIME_RANGES = {
    "5m": 5 * 60,
    "15m": 15 * 60,
    "30m": 30 * 60,
    "1h": 60 * 60,
    "2h": 2 * 60 * 60,
    "6h": 6 * 60 * 60,
    "12h": 12 * 60 * 60,
    "24h": 24 * 60 * 60,
    "2d": 2 * 24 * 60 * 60,
    "5d": 5 * 24 * 60 * 60,
    "1w": 7 * 24 * 60 * 60,
    "2w": 14 * 24 * 60 * 60,
    "1M": 30 * 24 * 60 * 60,
    "2M": 60 * 24 * 60 * 60,
    "3M": 90 * 24 * 60 * 60,
}
AUTHENTICATED_TIME_RANGES = frozenset({"2w", "1M", "2M", "3M"})

app = FastAPI(title="BrandMeister Statistics")
store: PostgresStore | None = None


def detect_locale(request: Request) -> str:
    """Match bm-lh-nextgen: cookie, query parameter, Accept-Language, English."""
    cookie_locale = normalize_locale(request.cookies.get(LANGUAGE_COOKIE))
    if cookie_locale:
        return cookie_locale
    query_locale = normalize_locale(request.query_params.get("lang"))
    if query_locale:
        return query_locale
    for language in request.headers.get("accept-language", "").split(","):
        locale = normalize_locale(language.split(";", 1)[0])
        if locale:
            return locale
    return "en"


def request_locale(request: Request) -> str:
    return getattr(request.state, "locale", None) or detect_locale(request)


@app.middleware("http")
async def language_middleware(request: Request, call_next: Any) -> Response:
    locale = detect_locale(request)
    request.state.locale = locale
    response = await call_next(request)
    if not normalize_locale(request.cookies.get(LANGUAGE_COOKIE)):
        query_locale = normalize_locale(request.query_params.get("lang"))
        if query_locale:
            response.set_cookie(
                LANGUAGE_COOKIE,
                query_locale,
                max_age=LANGUAGE_COOKIE_MAX_AGE,
                httponly=False,
                samesite="lax",
                secure=settings.cookie_secure,
            )
    response.headers["Content-Language"] = locale
    _refresh_user_session_cookie(request, response)
    return response


def get_store() -> PostgresStore:
    global store
    if store is None:
        store = PostgresStore(settings.database_url)
    return store


def start_time(time_range: str) -> datetime:
    return datetime.now(tz=UTC) - timedelta(seconds=TIME_RANGES.get(time_range, TIME_RANGES["5m"]))


def histogram_bucket_seconds(time_range: str) -> int:
    """Choose readable histogram bands for the selected dashboard period."""
    seconds = TIME_RANGES.get(time_range, TIME_RANGES["5m"])
    if seconds <= 15 * 60:
        return 60
    if seconds <= 2 * 60 * 60:
        return 5 * 60
    if seconds <= 12 * 60 * 60:
        return 30 * 60
    if seconds <= 24 * 60 * 60:
        return 60 * 60
    if seconds <= 7 * 24 * 60 * 60:
        return 6 * 60 * 60
    if seconds <= 31 * 24 * 60 * 60:
        return 24 * 60 * 60
    if seconds <= 62 * 24 * 60 * 60:
        return 3 * 24 * 60 * 60
    return 7 * 24 * 60 * 60


def _public_qso(row: dict[str, Any]) -> dict[str, Any]:
    return {
        "sessionId": row.get("session_id"),
        "sourceId": row.get("source_id"),
        "sourceCall": row.get("source_call"),
        "sourceName": row.get("source_name"),
        "destinationId": row.get("destination_id"),
        "destinationCall": row.get("destination_call"),
        "destinationName": row.get("destination_name"),
        "country": row.get("country"),
        "fullCountryName": row.get("full_country_name"),
        "continent": row.get("continent"),
        "contextId": row.get("context_id"),
        "linkCall": row.get("link_call"),
        "linkName": row.get("link_name"),
        "slot": row.get("slot"),
        "start": row.get("start_at"),
        "stop": row.get("stop_at"),
        "duration": (row.get("duration_ms") or 0) / 1000,
        "talkerAlias": row.get("talker_alias"),
    }


def _public_talkgroup(row: dict[str, Any]) -> dict[str, Any]:
    return {
        "destinationId": row.get("talkgroup_id"),
        "destinationName": row.get("destination_name"),
        "country": row.get("country"),
        "fullCountryName": row.get("full_country_name"),
        "continent": row.get("continent"),
        "count": row.get("qso_count", 0),
        "totalDuration": (row.get("total_duration_ms") or 0) / 1000,
        "uniqueSources": row.get("unique_sources", 0),
        "lastSeen": row.get("last_seen_at"),
    }


def _public_callsign(row: dict[str, Any]) -> dict[str, Any]:
    return {
        "callsign": row.get("callsign"),
        "sourceName": row.get("source_name"),
        "count": row.get("qso_count", 0),
        "totalDuration": (row.get("total_duration_ms") or 0) / 1000,
        "uniqueTalkgroups": row.get("unique_talkgroups", 0),
        "lastSeen": row.get("last_seen_at"),
    }


@app.on_event("startup")
def startup() -> None:
    get_store().initialize(settings.kerchunk_threshold_seconds)


@app.on_event("shutdown")
def shutdown() -> None:
    if store is not None:
        store.close()


@app.get("/health")
def health() -> dict[str, str]:
    return {"status": "ok"}


@app.get("/status")
def status() -> JSONResponse:
    """Expose database, collector, table, and active-user health metrics."""
    try:
        stale_after_seconds = max(settings.collector_heartbeat_seconds * 3, 90)
        payload = get_store().status_snapshot(stale_after_seconds)
    except Exception:
        return JSONResponse(
            {
                "status": "unhealthy",
                "database": {"status": "unhealthy", "connection": "failed"},
                "collector": {"status": "unknown"},
                "tables": {"raw_events": None, "qsos": None},
                "active_users": None,
            },
            status_code=503,
        )
    return JSONResponse(
        jsonable_encoder(payload),
        status_code=200 if payload["status"] == "ok" else 503,
    )


@app.get("/locales/{locale}")
def public_locale(locale: str) -> JSONResponse:
    normalized = normalize_locale(locale)
    if normalized not in SUPPORTED_LOCALES:
        return JSONResponse({"error": "unsupported locale"}, status_code=404)
    return JSONResponse(catalog(normalized))


@app.get("/api/stats/summary")
def stats_summary() -> dict[str, Any]:
    return get_store().summary()


@app.get("/api/qsos")
def qsos(
    request: Request,
    limit: int = Query(default=100, ge=1, le=500),
    offset: int = Query(default=0, ge=0),
    callsign: str | None = None,
    talkgroup: list[int] | None = Query(default=None),
) -> Any:
    access_error = _dashboard_access_error(request, callsign=callsign)
    if access_error is not None:
        return access_error
    return get_store().list_qsos(limit, offset, callsign, talkgroup)


@app.get("/public/stats")
def public_stats(
    request: Request,
    timeRange: str = "24h",
    continent: str | None = None,
    country: str | None = None,
    talkgroup: list[int] | None = Query(default=None),
    callsign: str | None = None,
) -> Any:
    access_error = _dashboard_access_error(request, time_range=timeRange, callsign=callsign)
    if access_error is not None:
        return access_error
    summary = get_store().summary(start_time(timeRange), continent, country, talkgroup, callsign)
    histogram = get_store().activity_histogram(
        start_time(timeRange),
        histogram_bucket_seconds(timeRange),
        continent,
        country,
        talkgroup,
        callsign,
    )
    return {
        "totalEntries": summary["qso_count"],
        "activityRange": summary["qso_count"],
        "activity24h": summary["qso_count"],
        "uniqueCallsigns": summary["unique_sources"],
        "uniqueTalkgroups": summary["unique_destinations"],
        "totalDuration": summary["duration_seconds"],
        "durationRange": summary["duration_seconds"],
        "firstQsoAt": summary["first_qso_at"],
        "lastQsoAt": summary["last_qso_at"],
        "timeRange": timeRange,
        "histogram": histogram,
    }


@app.get("/public/lastheard")
def public_lastheard(
    request: Request,
    limit: int = Query(default=50, ge=1, le=500),
    callsign: str | None = None,
    talkgroup: list[int] | None = Query(default=None),
) -> Any:
    access_error = _dashboard_access_error(request, callsign=callsign)
    if access_error is not None:
        return access_error
    rows = get_store().list_qsos(
        limit, 0, callsign, talkgroup,
        min_duration_seconds=settings.kerchunk_threshold_seconds,
    )
    return [_public_qso(row) for row in rows]


@app.get("/public/lastheard/grouped")
def public_grouped_lastheard(
    request: Request,
    timeRange: str = "5m",
    limit: int = Query(default=25, ge=1, le=50),
    continent: str | None = None,
    country: str | None = None,
    talkgroup: list[int] | None = Query(default=None),
    callsign: str | None = None,
) -> Any:
    access_error = _dashboard_access_error(request, time_range=timeRange, callsign=callsign)
    if access_error is not None:
        return access_error
    rows = get_store().grouped_by_talkgroup(
        start_time(timeRange), limit, continent, country, talkgroup, callsign
    )
    return [_public_talkgroup(row) for row in rows]


@app.get("/public/lastheard/callsigns")
def public_grouped_callsigns(
    request: Request,
    timeRange: str = "5m",
    limit: int = Query(default=25, ge=1, le=50),
    callsign: str | None = None,
    continent: str | None = None,
    country: str | None = None,
    talkgroup: list[int] | None = Query(default=None),
) -> Any:
    access_error = _dashboard_access_error(
        request, callsign=callsign, time_range=timeRange
    )
    if access_error is not None:
        return access_error
    rows = get_store().grouped_by_callsign(
        start_time(timeRange), limit, callsign, continent, country, talkgroup
    )
    return [_public_callsign(row) for row in rows]


@app.get("/public/continents")
def public_continents() -> list[str]:
    return get_store().continents()


@app.get("/public/countries")
def public_countries(continent: str | None = None) -> list[dict[str, str]]:
    return get_store().countries(continent)


@app.get("/public/talkgroups")
def public_talkgroups(
    continent: str | None = None,
    country: str | None = None,
) -> list[dict[str, Any]]:
    return get_store().talkgroups(continent, country)


@app.get("/user/talkgroups")
def active_user_talkgroups(
    request: Request,
    timeRange: str = "30m",
    continent: str | None = None,
    country: str | None = None,
) -> Any:
    if _current_user(request) is None:
        return JSONResponse({"error": "authentication required"}, status_code=401)
    access_error = _dashboard_access_error(request, time_range=timeRange)
    if access_error is not None:
        return access_error
    rows = get_store().active_talkgroups(start_time(timeRange), continent, country)
    return [
        {
            "value": row["value"],
            "label": row["label"],
            "count": row["count"],
            "totalDuration": row["total_duration_ms"] / 1000,
        }
        for row in rows
    ]


def _index_path() -> Path:
    candidates = (
        Path(__file__).resolve().parents[2] / "static" / "index.html",
        Path.cwd() / "static" / "index.html",
        Path("/app/static/index.html"),
    )
    for candidate in candidates:
        if candidate.exists():
            return candidate
    raise FileNotFoundError("static/index.html is not available")


def _cookie_labels(locale: str) -> dict[str, str]:
    return {
        key: translate(locale, f"cookies.{key}")
        for key in (
            "title",
            "description",
            "acceptAnalytics",
            "rejectAnalytics",
            "settings",
            "settingsTitle",
            "necessary",
            "necessaryDescription",
            "analytics",
            "analyticsDescription",
            "save",
            "continue",
        )
    }


@app.get("/")
def dashboard(request: Request = None) -> HTMLResponse:
    page = _index_path().read_text(encoding="utf-8")
    user = _current_user(request) if request is not None else None
    locale = request_locale(request) if request is not None else "en"
    extended_range_options = (
        '<option value="2w" data-auth-required>Last 14 days</option>'
        '<option value="1M" data-auth-required>Last month</option>'
        '<option value="2M" data-auth-required>Last 2 months</option>'
        '<option value="3M" data-auth-required>Last 3 months</option>'
        if user is not None
        else ""
    )
    callsign_search = (
        '<div class="control search"><label for="callsign" '
        'data-i18n="home.callsignFilter">Callsign filter</label>'
        '<input id="callsign" data-i18n-placeholder="home.callsignPlaceholder" '
        'placeholder="e.g. EA7KLK"></div>'
        if user is not None
        else ""
    )
    talkgroup_filter = (
        '<div class="control talkgroup-filter" id="talkgroupControl" style="display:none">'
        '<label for="talkgroups" data-i18n="home.talkgroupFilter">Talkgroups</label>'
        '<select id="talkgroups" multiple size="5" aria-describedby="talkgroupFilterHint"></select>'
        '<small id="talkgroupFilterHint" data-i18n="home.talkgroupFilterHint">Select one or more active talkgroups</small></div>'
        if user is not None
        else ""
    )
    page = page.replace("<!-- AUTHENTICATED_CALLSIGN_SEARCH -->", callsign_search, 1)
    page = page.replace("<!-- AUTHENTICATED_TALKGROUP_FILTER -->", talkgroup_filter, 1)
    page = page.replace(
        "<!-- AUTHENTICATED_TIME_RANGES -->", extended_range_options, 1
    )
    page = page.replace(
        '<main class="shell" data-authenticated="false">',
        f'<main class="shell" data-authenticated="{str(user is not None).lower()}">',
        1,
    )
    analytics_enabled = matomo_configured()
    page = page.replace(
        "<!-- COOKIE_CONSENT_MARKUP -->",
        cookie_consent_markup(_cookie_labels(locale), analytics_enabled),
        1,
    )
    if user is not None:
        page = page.replace(
            '<a href="/user/login" data-i18n="home.login">Log in</a>',
            "",
            1,
        )
        page = page.replace(
            '<a href="/user/register" data-i18n="home.register">Register</a>',
            "",
            1,
        )
        page = page.replace(
            '<a href="/user/profile" data-i18n="home.myProfile">My profile</a>',
            f'<a href="/user/profile" data-user-callsign>{_escape(user["callsign"])}</a>'
            '<a href="/user/live-qsos" data-i18n="live.title">Live QSOs</a>'
            '<a href="/user/reports" data-i18n="home.reports">Reports</a>'
            '<form class="account-logout" method="post" action="/user/logout">'
            '<button type="submit" data-i18n="user.logout">Log out</button></form>',
            1,
        )
    page = page.replace("</head>", f"{matomo_script()}\n</head>", 1)
    page = page.replace(
        "<!-- COOKIE_CONSENT_SCRIPT -->",
        cookie_consent_script(analytics_enabled),
        1,
    )
    return HTMLResponse(page)


@app.get("/about", response_class=HTMLResponse)
def about_page(request: Request) -> HTMLResponse:
    locale = request_locale(request)
    content = f"""
<section class="card">
  <h1>{_escape(translate(locale, "about.title"))}</h1>
  <p class="muted" style="font-size:16px;line-height:1.7">{_escape(translate(locale, "about.intro"))}</p>
  <p style="padding:14px 16px;border-radius:10px;background:#fff7ed;color:#9a3412;line-height:1.6">{_escape(translate(locale, "about.accountMigration"))}</p>
  <div style="display:grid;grid-template-columns:repeat(auto-fit,minmax(220px,1fr));gap:16px;margin-top:24px">
    <div style="padding:18px;border-radius:10px;background:#f8f9ff">
      <h2 style="font-size:17px">{_escape(translate(locale, "about.project"))}</h2>
      <p class="muted">BrandMeister Lastheard</p>
      <p><a href="https://github.com/ea7klk/bm-lh-new" target="_blank" rel="noopener noreferrer">{_escape(translate(locale, "about.githubLink"))}</a></p>
      <p class="muted">{_escape(translate(locale, "about.repositoryText"))}</p>
    </div>
    <div style="padding:18px;border-radius:10px;background:#f8f9ff">
      <h2 style="font-size:17px">{_escape(translate(locale, "about.author"))}</h2>
      <p>{_escape(translate(locale, "about.authorName"))}</p>
      <h2 style="font-size:17px;margin-top:22px">{_escape(translate(locale, "about.copyright"))}</h2>
      <p>© 2026 Volker Kerkhoff</p>
    </div>
    <div style="padding:18px;border-radius:10px;background:#f8f9ff">
      <h2 style="font-size:17px">{_escape(translate(locale, "about.license"))}</h2>
      <p class="muted">{_escape(translate(locale, "about.licenseText"))}</p>
      <p><a href="https://creativecommons.org/licenses/by-nc-sa/4.0/" target="_blank" rel="noopener noreferrer">CC BY-NC-SA 4.0</a></p>
    </div>
    <div style="padding:18px;border-radius:10px;background:#f8f9ff">
      <h2 style="font-size:17px">{_escape(translate(locale, "about.cookieTitle"))}</h2>
      <p class="muted">{_escape(translate(locale, "about.cookieText"))}</p>
    </div>
  </div>
  <p style="margin-top:26px"><a class="button secondary" href="/">{_escape(translate(locale, "about.back"))}</a></p>
</section>
"""
    return _account_page(translate(locale, "about.title"), content, locale)


def _escape(value: Any) -> str:
    return html.escape("" if value is None else str(value))


def _format_datetime(value: Any) -> str:
    return "—" if value is None else str(value).replace("+00:00", " UTC")


def _account_page(title: str, content: str, locale: str = "en") -> HTMLResponse:
    return _account_page_with_metrics(title, content, locale)


def _account_page_with_metrics(
    title: str,
    content: str,
    locale: str = "en",
    records_retrieved: int = 0,
    query_seconds: float = 0.0,
) -> HTMLResponse:
    locale = normalize_locale(locale) or "en"
    language_options = "".join(
        f'<option value="{code}"{" selected" if code == locale else ""}>{info["flag"]} {info["name"]}</option>'
        for code, info in LANGUAGE_INFO.items()
    )
    analytics_enabled = matomo_configured()
    consent_markup = (
        cookie_consent_markup(_cookie_labels(locale), analytics_enabled)
    )
    metrics = translate(locale, "common.recordsRetrieved").format(
        records=int(records_retrieved),
        seconds=f"{max(float(query_seconds), 0.0):.3f}",
    )
    return HTMLResponse(
        f"""
<!doctype html><html lang="{locale}"><head><meta charset="utf-8"><meta name="viewport" content="width=device-width,initial-scale=1">
<title>{_escape(title)} · BrandMeister</title>
{matomo_script()}
<style>
body{{margin:0;min-height:100vh;padding:24px;background:linear-gradient(135deg,#667eea,#764ba2);font-family:system-ui,-apple-system,"Segoe UI",sans-serif;color:#1f2937}}
.shell{{max-width:1100px;margin:auto}}.card{{background:#fff;border-radius:14px;box-shadow:0 18px 55px #1e153a33;padding:28px;margin-bottom:20px}}
h1,h2{{margin-top:0}}h1{{text-align:center}}.muted{{color:#6b7280}}.nav{{display:flex;justify-content:space-between;align-items:center;margin-bottom:18px;gap:12px;flex-wrap:wrap}}
.nav a,.button{{display:inline-block;padding:9px 14px;border-radius:8px;border:0;background:linear-gradient(135deg,#667eea,#764ba2);color:white;text-decoration:none;font-weight:700;cursor:pointer}}
.nav a.secondary,.button.secondary{{background:#f1f3ff;color:#5457bd}}.language{{display:flex;align-items:center;gap:7px;color:#fff;font-size:13px;font-weight:700}}.language select{{padding:7px 9px;border:0;border-radius:7px;background:#fff;color:#374151;font:inherit}}.form{{max-width:520px;margin:auto}}label{{display:block;margin:13px 0 5px;font-size:13px;font-weight:700;color:#4b5563}}input{{width:100%;height:42px;padding:0 11px;border:2px solid #e5e7eb;border-radius:8px;box-sizing:border-box;font:inherit}}input:focus{{outline:0;border-color:#667eea}}.form .button{{margin-top:18px;width:100%}}
.error{{padding:12px;border-radius:8px;background:#fff1f2;color:#be123c;margin:0 0 15px}}.success{{padding:12px;border-radius:8px;background:#ecfdf3;color:#15803d;margin:0 0 15px}}
.stats{{display:grid;grid-template-columns:repeat(4,1fr);gap:14px}}.stat{{background:#f8f9ff;border-radius:10px;padding:16px}}.stat small{{color:#6b7280;text-transform:uppercase;font-weight:800;letter-spacing:.05em}}.stat strong{{display:block;font-size:25px;margin-top:7px}}
.charts{{display:grid;grid-template-columns:1fr 1fr;gap:20px}}.charts h3{{margin:0 0 10px;font-size:15px}}
table{{width:100%;border-collapse:collapse}}th,td{{padding:11px;border-bottom:1px solid #e8eaf0;text-align:left;font-size:13px}}th{{background:linear-gradient(135deg,#667eea,#764ba2);color:#fff}}.table-wrap{{overflow:auto;border:1px solid #e8eaf0;border-radius:9px}}.inline{{display:inline}}.danger{{background:#dc3545}}.warning{{color:#b45309;font-weight:700}}
.cookie-consent{{position:fixed;z-index:1000;left:16px;right:16px;bottom:16px;display:flex;justify-content:center}}.cookie-consent[hidden]{{display:none}}.cookie-consent-card{{max-width:760px;width:100%;padding:20px 22px;background:#fff;border:1px solid #e5e7eb;border-radius:14px;box-shadow:0 18px 55px #1e153a55}}.cookie-consent-card h2,.cookie-consent-card h3{{margin:0 0 8px}}.cookie-consent-card p{{margin:0;color:#4b5563;line-height:1.5;font-size:14px}}.cookie-consent-actions{{display:flex;align-items:center;gap:9px;flex-wrap:wrap;margin-top:15px}}.cookie-consent-actions .button{{width:auto;margin:0}}.cookie-settings-link,.cookie-settings-footer{{border:0;background:none;color:#5457bd;text-decoration:underline;cursor:pointer;font:inherit;font-size:13px}}.cookie-settings{{margin-top:15px;padding-top:15px;border-top:1px solid #e5e7eb}}.cookie-option{{display:flex;align-items:flex-start;gap:9px;margin:11px 0;font-weight:400}}.cookie-option input{{width:auto;height:auto;margin-top:3px}}.cookie-option span{{display:flex;flex-direction:column;gap:3px}}.cookie-option small{{color:#6b7280;font-weight:400;line-height:1.4}}.cookie-settings-footer{{display:block;margin:24px auto 0;color:#fff}}.page-footer{{color:#fff;text-align:center;font-size:12px;line-height:1.6;padding:4px 0 8px}}.page-footer a{{color:#fff}}
@media(max-width:700px){{body{{padding:10px}}.card{{padding:20px}}.stats{{grid-template-columns:repeat(2,1fr)}}.charts{{grid-template-columns:1fr}}table{{min-width:760px}}}}
</style></head><body><main class="shell"><div class="nav"><a href="/">{_escape(translate(locale, "common.dashboard"))}</a><span style="color:white;font-weight:800">{_escape(translate(locale, "common.accounts"))}</span><label class="language">{_escape(translate(locale, "common.language"))}<select id="language">{language_options}</select></label></div>{content}{consent_markup}<footer class="page-footer"><span>{_escape(metrics)}</span> · <a href="/">{_escape(translate(locale, "common.dashboard"))}</a></footer></main><script>document.getElementById('language').addEventListener('change',function(){{document.cookie='{LANGUAGE_COOKIE}='+encodeURIComponent(this.value)+'; Max-Age={LANGUAGE_COOKIE_MAX_AGE}; Path=/; SameSite=Lax';window.location.reload();}});</script>{cookie_consent_script(analytics_enabled)}</body></html>
"""
    )


async def _form_fields(request: Request) -> dict[str, str]:
    body = (await request.body()).decode("utf-8", errors="replace")
    if request.headers.get("content-type", "").startswith("application/json"):
        try:
            data = await request.json()
            return {str(key): str(value) for key, value in data.items()}
        except ValueError:
            return {}
    return {key: values[-1] for key, values in parse_qs(body).items()}


def _current_user(request: Request) -> dict[str, Any] | None:
    request.state.user_session_checked = True
    token = request.cookies.get("session_token")
    if not token:
        request.state.user_session_valid = False
        return None
    user = get_store().user_by_session(
        session_token_hash(token), settings.session_hours * 60 * 60
    )
    request.state.user_session_valid = user is not None
    return user


def _dashboard_access_error(
    request: Request,
    *,
    callsign: str | None = None,
    time_range: str | None = None,
) -> JSONResponse | None:
    """Require a signed-in user for callsign searches and extended ranges."""
    requires_authentication = bool(callsign and callsign.strip()) or (
        time_range in AUTHENTICATED_TIME_RANGES
    )
    if requires_authentication and _current_user(request) is None:
        return JSONResponse({"error": "authentication required"}, status_code=401)
    return None


def _refresh_user_session_cookie(request: Request, response: Response) -> None:
    """Keep the browser cookie aligned with the sliding database expiry."""
    if not getattr(request.state, "user_session_checked", False):
        return
    token = request.cookies.get("session_token")
    if not token:
        return
    if getattr(request.state, "user_session_valid", False):
        response.set_cookie(
            "session_token",
            token,
            max_age=settings.session_hours * 3600,
            httponly=True,
            samesite="lax",
            secure=settings.cookie_secure,
        )
    else:
        response.delete_cookie("session_token")


def _user_redirect(user_id: int) -> RedirectResponse:
    token = new_session_token()
    get_store().create_session(
        user_id,
        session_token_hash(token),
        datetime.now(tz=UTC) + timedelta(hours=settings.session_hours),
    )
    response = RedirectResponse("/user/profile", status_code=303)
    response.set_cookie(
        "session_token", token, max_age=settings.session_hours * 3600,
        httponly=True, samesite="lax", secure=settings.cookie_secure,
    )
    return response


def _admin_allowed(request: Request) -> bool:
    return verify_admin_token(request.cookies.get("admin_session"), settings.admin_password)


def _admin_redirect(query: str = "") -> RedirectResponse:
    suffix = f"?{query}" if query else ""
    return RedirectResponse(f"/admin{suffix}", status_code=303)


def _validation_error(message: str, path: str) -> HTMLResponse:
    return _account_page(path, f'<div class="card"><p class="error">{_escape(message)}</p><p><a class="button" href="/user/{path.lower()}">Back</a></p></div>')


@app.get("/user/register", response_class=HTMLResponse)
def user_register_page(request: Request) -> Response:
    locale = request_locale(request)
    return _account_page(
        translate(locale, "user.register"),
        f"""
<section class="card form"><h1>{_escape(translate(locale, "user.createAccount"))}</h1><p class="muted">{_escape(translate(locale, "user.registerPrompt"))}</p>
<form method="post" action="/user/register"><label for="callsign">{_escape(translate(locale, "user.callsign"))}</label><input id="callsign" name="callsign" required maxlength="16" autocomplete="username">
<label for="name">{_escape(translate(locale, "user.name"))}</label><input id="name" name="name" required maxlength="120" autocomplete="name">
<label for="email">{_escape(translate(locale, "user.email"))}</label><input id="email" type="email" name="email" required maxlength="240" autocomplete="email">
<label for="password">{_escape(translate(locale, "user.password"))}</label><input id="password" type="password" name="password" required minlength="8" autocomplete="new-password">
<button class="button" type="submit">{_escape(translate(locale, "user.createAccountButton"))}</button></form><p class="muted">{_escape(translate(locale, "user.alreadyRegistered"))} <a href="/user/login">{_escape(translate(locale, "user.login"))}</a>.</p></section>
""",
        locale,
    )


@app.post("/user/register")
async def user_register(request: Request) -> Response:
    locale = request_locale(request)
    fields = await _form_fields(request)
    callsign = fields.get("callsign", "").strip().upper()
    name = fields.get("name", "").strip()
    email = fields.get("email", "").strip().lower()
    password = fields.get("password", "")
    if not re.fullmatch(r"[A-Z0-9][A-Z0-9/-]{1,15}", callsign):
        return _account_page(translate(locale, "user.register"), f'<section class="card"><p class="error">{_escape(translate(locale, "user.invalidCallsign"))}</p><a class="button" href="/user/register">{_escape(translate(locale, "common.dashboard"))}</a></section>', locale)
    if not name or len(name) > 120 or not re.fullmatch(r"[^@\s]+@[^@\s]+\.[^@\s]+", email):
        return _account_page(translate(locale, "user.register"), f'<section class="card"><p class="error">{_escape(translate(locale, "user.invalidNameEmail"))}</p><a class="button" href="/user/register">{_escape(translate(locale, "common.dashboard"))}</a></section>', locale)
    if len(password) < 8:
        return _account_page(translate(locale, "user.register"), f'<section class="card"><p class="error">{_escape(translate(locale, "user.passwordMin"))}</p><a class="button" href="/user/register">{_escape(translate(locale, "common.dashboard"))}</a></section>', locale)
    try:
        user = get_store().create_user(callsign, name, email, hash_password(password))
    except UniqueViolation:
        return _account_page(translate(locale, "user.register"), f'<section class="card"><p class="error">{_escape(translate(locale, "user.duplicate"))}</p><a class="button" href="/user/register">{_escape(translate(locale, "common.dashboard"))}</a></section>', locale)
    token = new_email_verification_token()
    try:
        get_store().create_email_verification(
            user["id"],
            session_token_hash(token),
            datetime.now(tz=UTC) + timedelta(hours=settings.email_verification_hours),
        )
        send_verification_email(email, callsign, locale, token)
    except EmailDeliveryError:
        get_store().delete_user(user["id"])
        return _account_page(
            translate(locale, "user.register"),
            f'<section class="card"><h1>{_escape(translate(locale, "emailVerification.invalidTitle"))}</h1><p class="error">{_escape(translate(locale, "emailVerification.deliveryFailed"))}</p><a class="button" href="/user/register">{_escape(translate(locale, "user.register"))}</a></section>',
            locale,
        )
    return _account_page(
        translate(locale, "emailVerification.sentTitle"),
        f'<section class="card form"><h1>{_escape(translate(locale, "emailVerification.sentTitle"))}</h1><p class="success">{_escape(translate(locale, "emailVerification.sent"))}</p><a class="button" href="/user/login">{_escape(translate(locale, "emailVerification.login"))}</a></section>',
        locale,
    )


@app.get("/user/verify", response_class=HTMLResponse)
def user_verify(request: Request, token: str = "") -> Response:
    locale = request_locale(request)
    user = get_store().verify_email_token(session_token_hash(token)) if token else None
    if user is None:
        return _account_page(
            translate(locale, "emailVerification.invalidTitle"),
            f'<section class="card form"><h1>{_escape(translate(locale, "emailVerification.invalidTitle"))}</h1><p class="error">{_escape(translate(locale, "emailVerification.invalid"))}</p><a class="button" href="/user/register">{_escape(translate(locale, "user.register"))}</a></section>',
            locale,
        )
    return _account_page(
        translate(locale, "emailVerification.successTitle"),
        f'<section class="card form"><h1>{_escape(translate(locale, "emailVerification.successTitle"))}</h1><p class="success">{_escape(translate(locale, "emailVerification.success"))}</p><a class="button" href="/user/login">{_escape(translate(locale, "emailVerification.login"))}</a></section>',
        locale,
    )


def _password_reset_invalid_page(locale: str) -> HTMLResponse:
    return _account_page(
        translate(locale, "user.resetInvalidTitle"),
        f'<section class="card form"><h1>{_escape(translate(locale, "user.resetInvalidTitle"))}</h1><p class="error">{_escape(translate(locale, "user.resetInvalid"))}</p><a class="button" href="/user/forgot-password">{_escape(translate(locale, "user.forgotPassword"))}</a></section>',
        locale,
    )


def _password_reset_sent_page(locale: str) -> HTMLResponse:
    return _account_page(
        translate(locale, "user.resetEmailSentTitle"),
        f'<section class="card form"><h1>{_escape(translate(locale, "user.resetEmailSentTitle"))}</h1><p class="success">{_escape(translate(locale, "user.resetEmailSent"))}</p><a class="button" href="/user/login">{_escape(translate(locale, "user.passwordResetLogin"))}</a></section>',
        locale,
    )


@app.get("/user/forgot-password", response_class=HTMLResponse)
def user_forgot_password_page(request: Request) -> HTMLResponse:
    locale = request_locale(request)
    return _account_page(
        translate(locale, "user.forgotPasswordTitle"),
        f'<section class="card form"><h1>{_escape(translate(locale, "user.forgotPasswordTitle"))}</h1><p class="muted">{_escape(translate(locale, "user.forgotPasswordPrompt"))}</p><form method="post" action="/user/forgot-password"><label for="email">{_escape(translate(locale, "user.email"))}</label><input id="email" type="email" name="email" required maxlength="240" autocomplete="email"><button class="button" type="submit">{_escape(translate(locale, "user.sendResetLink"))}</button></form><p class="muted"><a href="/user/login">{_escape(translate(locale, "user.login"))}</a></p></section>',
        locale,
    )


@app.post("/user/forgot-password")
async def user_forgot_password(request: Request) -> Response:
    locale = request_locale(request)
    fields = await _form_fields(request)
    email = fields.get("email", "").strip().lower()
    user = get_store().user_by_email(email) if re.fullmatch(r"[^@\s]+@[^@\s]+\.[^@\s]+", email) else None
    if user is not None:
        token = new_password_reset_token()
        get_store().create_password_reset(
            user["id"],
            session_token_hash(token),
            datetime.now(tz=UTC) + timedelta(hours=settings.password_reset_hours),
        )
        try:
            send_password_reset_email(user["email"], user["callsign"], locale, token)
        except EmailDeliveryError:
            return _account_page(
                translate(locale, "user.forgotPasswordTitle"),
                f'<section class="card form"><h1>{_escape(translate(locale, "user.forgotPasswordTitle"))}</h1><p class="error">{_escape(translate(locale, "user.passwordResetDeliveryFailed"))}</p><a class="button" href="/user/forgot-password">{_escape(translate(locale, "user.forgotPassword"))}</a></section>',
                locale,
            )
    return _password_reset_sent_page(locale)


@app.get("/user/reset-password", response_class=HTMLResponse)
def user_reset_password_page(request: Request, token: str = "") -> Response:
    locale = request_locale(request)
    token_hash = session_token_hash(token) if token else ""
    if not token_hash or get_store().password_reset_user(token_hash) is None:
        return _password_reset_invalid_page(locale)
    return _account_page(
        translate(locale, "user.resetPasswordTitle"),
        f'<section class="card form"><h1>{_escape(translate(locale, "user.resetPasswordTitle"))}</h1><p class="muted">{_escape(translate(locale, "user.resetPasswordPrompt"))}</p><form method="post" action="/user/reset-password"><input type="hidden" name="token" value="{_escape(token)}"><label for="password">{_escape(translate(locale, "user.newPassword"))}</label><input id="password" type="password" name="password" minlength="8" required autocomplete="new-password"><label for="confirm_password">{_escape(translate(locale, "user.confirmPassword"))}</label><input id="confirm_password" type="password" name="confirm_password" minlength="8" required autocomplete="new-password"><button class="button" type="submit">{_escape(translate(locale, "user.resetPasswordButton"))}</button></form></section>',
        locale,
    )


@app.post("/user/reset-password")
async def user_reset_password(request: Request) -> Response:
    locale = request_locale(request)
    fields = await _form_fields(request)
    token = fields.get("token", "")
    password = fields.get("password", "")
    if len(password) < 8:
        return _account_page(
            translate(locale, "user.resetPasswordTitle"),
            f'<section class="card form"><h1>{_escape(translate(locale, "user.resetPasswordTitle"))}</h1><p class="error">{_escape(translate(locale, "user.passwordMin"))}</p><a class="button" href="/user/reset-password?token={_escape(quote(token))}">{_escape(translate(locale, "user.resetPasswordTitle"))}</a></section>',
            locale,
        )
    if password != fields.get("confirm_password", ""):
        return _account_page(
            translate(locale, "user.resetPasswordTitle"),
            f'<section class="card form"><h1>{_escape(translate(locale, "user.resetPasswordTitle"))}</h1><p class="error">{_escape(translate(locale, "user.passwordMismatch"))}</p><a class="button" href="/user/reset-password?token={_escape(quote(token))}">{_escape(translate(locale, "user.resetPasswordTitle"))}</a></section>',
            locale,
        )
    user = get_store().reset_password(session_token_hash(token), hash_password(password)) if token else None
    if user is None:
        return _password_reset_invalid_page(locale)
    return _account_page(
        translate(locale, "user.resetPasswordTitle"),
        f'<section class="card form"><h1>{_escape(translate(locale, "user.resetPasswordTitle"))}</h1><p class="success">{_escape(translate(locale, "user.passwordResetSuccess"))}</p><a class="button" href="/user/login">{_escape(translate(locale, "user.passwordResetLogin"))}</a></section>',
        locale,
    )


@app.get("/user/login", response_class=HTMLResponse)
def user_login_page(request: Request) -> HTMLResponse:
    locale = request_locale(request)
    return _account_page(
        translate(locale, "user.loginTitle"),
        f"""
<section class="card form"><h1>{_escape(translate(locale, "user.loginTitle"))}</h1><p class="muted">{_escape(translate(locale, "user.loginPrompt"))}</p>
<form method="post" action="/user/login"><label for="login">{_escape(translate(locale, "user.emailOrCallsign"))}</label><input id="login" name="login" required autocomplete="username">
<label for="password">{_escape(translate(locale, "user.password"))}</label><input id="password" type="password" name="password" required autocomplete="current-password">
<button class="button" type="submit">{_escape(translate(locale, "user.loginButton"))}</button></form><p class="muted"><a href="/user/forgot-password">{_escape(translate(locale, "user.forgotPassword"))}</a></p><p class="muted">{_escape(translate(locale, "user.needAccount"))} <a href="/user/register">{_escape(translate(locale, "user.register"))}</a>.</p></section>
""",
        locale,
    )


@app.post("/user/login")
async def user_login(request: Request) -> Response:
    locale = request_locale(request)
    fields = await _form_fields(request)
    login = (fields.get("login") or fields.get("email") or fields.get("callsign") or "").strip()
    user = get_store().user_by_login(login)
    if user is None or not verify_password(fields.get("password", ""), user["password_hash"]):
        return _account_page(translate(locale, "user.loginTitle"), f'<section class="card"><p class="error">{_escape(translate(locale, "user.invalidCredentials"))}</p><p class="muted">{_escape(translate(locale, "user.accountMayNeedRegistration"))} <a href="/user/register">{_escape(translate(locale, "user.register"))}</a></p><p><a class="button" href="/user/forgot-password">{_escape(translate(locale, "user.forgotPassword"))}</a></p><a class="button secondary" href="/user/login">{_escape(translate(locale, "user.loginButton"))}</a></section>', locale)
    if not user["is_active"]:
        return _account_page(translate(locale, "user.loginTitle"), f'<section class="card"><p class="error">{_escape(translate(locale, "emailVerification.notVerified"))}</p><a class="button" href="/user/login">{_escape(translate(locale, "user.loginButton"))}</a></section>', locale)
    get_store().mark_user_login(user["id"])
    if is_bcrypt_hash(user["password_hash"]):
        get_store().update_user_password(user["id"], hash_password(fields.get("password", "")))
    return _user_redirect(user["id"])


@app.post("/user/logout")
def user_logout(request: Request) -> RedirectResponse:
    token = request.cookies.get("session_token")
    if token:
        get_store().delete_session(session_token_hash(token))
    response = RedirectResponse("/", status_code=303)
    response.delete_cookie("session_token")
    return response


@app.get("/user/profile", response_class=HTMLResponse)
def user_profile(request: Request) -> Response:
    locale = request_locale(request)
    user = _current_user(request)
    if user is None:
        return RedirectResponse("/user/login", status_code=303)
    query_started = perf_counter()
    stats = get_store().user_statistics(user["callsign"])
    query_seconds = perf_counter() - query_started
    top_rows = "".join(
        f"<tr><td>{_escape(row['name'])}</td><td>{_escape(row['talkgroup_id'])}</td><td>{_escape(row['qso_count'])}</td><td>{_escape(round(row['duration_seconds'], 1))} s</td></tr>"
        for row in stats["top_talkgroups"]
    ) or f'<tr><td colspan="4" class="muted">{_escape(translate(locale, "user.noQsos"))}</td></tr>'
    content = f"""
<section class="card"><div class="nav"><h1 style="margin:0">{_escape(user['callsign'])}</h1><div><a class="button secondary" href="/user/live-qsos">{_escape(translate(locale, "live.title"))}</a> <a class="button secondary" href="/user/reports">{_escape(translate(locale, "home.reports"))}</a> <form class="inline" method="post" action="/user/logout"><button class="button secondary" type="submit">{_escape(translate(locale, "user.logout"))}</button></form></div></div><p class="muted">{_escape(user['name'])} · {_escape(user['email'])}</p>
<div class="stats"><div class="stat"><small>{_escape(translate(locale, "user.qsoCount"))}</small><strong>{_escape(stats['qso_count'])}</strong></div><div class="stat"><small>{_escape(translate(locale, "home.talkTime"))}</small><strong>{_escape(round(stats['duration_seconds'],1))} s</strong></div><div class="stat"><small>{_escape(translate(locale, "user.uniqueTalkgroups"))}</small><strong>{_escape(stats['unique_talkgroups'])}</strong></div><div class="stat"><small>{_escape(translate(locale, "home.lastHeard"))}</small><strong>{_escape(_format_datetime(stats['last_qso_at']))}</strong></div></div></section>
<section class="card"><h2>{_escape(translate(locale, "user.topTalkgroups"))}</h2><div class="table-wrap"><table><thead><tr><th>{_escape(translate(locale, "home.talkgroup"))}</th><th>{_escape(translate(locale, "user.id"))}</th><th>{_escape(translate(locale, "user.qsoCount"))}</th><th>{_escape(translate(locale, "home.talkTime"))}</th></tr></thead><tbody>{top_rows}</tbody></table></div></section>
<section class="card form"><h2>{_escape(translate(locale, "user.changePassword"))}</h2><form method="post" action="/user/change-password"><label>{_escape(translate(locale, "user.currentPassword"))}</label><input type="password" name="current_password" required><label>{_escape(translate(locale, "user.newPassword"))}</label><input type="password" name="new_password" minlength="8" required><button class="button" type="submit">{_escape(translate(locale, "user.changePasswordButton"))}</button></form></section>
"""
    return _account_page_with_metrics(
        translate(locale, "user.profile"),
        content,
        locale,
        records_retrieved=stats["qso_count"],
        query_seconds=query_seconds,
    )


def _live_time_range(value: str | None) -> str:
    return value if value in TIME_RANGES else "30m"


def _relative_time(value: datetime | None, locale: str) -> str:
    if value is None:
        return "—"
    if value.tzinfo is None:
        value = value.replace(tzinfo=UTC)
    elapsed = max(0, int((datetime.now(tz=UTC) - value).total_seconds()))
    if elapsed < 5:
        return translate(locale, "live.justNow")
    if elapsed < 60:
        amount, unit = elapsed, "secondsUnit"
    elif elapsed < 3600:
        amount, unit = elapsed // 60, "minutesUnit"
    elif elapsed < 86400:
        amount, unit = elapsed // 3600, "hoursUnit"
    else:
        amount, unit = elapsed // 86400, "daysUnit"
    prefix = translate(locale, "live.relativePrefix")
    suffix = translate(locale, "live.relativeSuffix")
    return f"{prefix}{amount} {translate(locale, f'live.{unit}')}{suffix}"


def _live_qso_rows(rows: list[dict[str, Any]], locale: str) -> str:
    if not rows:
        return f'<tr><td colspan="5" class="muted">{_escape(translate(locale, "live.noData"))}</td></tr>'
    return "".join(
        f'<tr><td>{_escape(_relative_time(row["start_at"], locale))}</td>'
        f'<td><span class="live-primary">{_escape(row.get("source_call") or row.get("source_id") or "—")}</span>'
        f' <span class="live-muted">{_escape(row.get("source_name") or "")}</span></td>'
        f'<td><span class="live-primary">{_escape(row.get("destination_name") or "—")}</span>'
        f' <span class="live-muted">({_escape(row.get("destination_id") or "—")})</span></td>'
        f'<td>{_escape(row.get("slot") or "—")}</td>'
        f'<td class="live-duration">{_escape(_report_duration((row.get("duration_ms") or 0) / 1000))}</td></tr>'
        for row in rows
    )


def _live_subscription(data: Any) -> dict[str, Any]:
    data = data if isinstance(data, dict) else {}
    raw_talkgroups = data.get("talkgroups", data.get("talkgroup", []))
    if not isinstance(raw_talkgroups, list):
        raw_talkgroups = [raw_talkgroups]
    talkgroups: set[int] = set()
    for value in raw_talkgroups:
        try:
            talkgroups.add(int(value))
        except (TypeError, ValueError):
            continue
    try:
        rows = int(data.get("rows", 25))
    except (TypeError, ValueError):
        rows = 25
    return {
        "time_range": _live_time_range(str(data.get("timeRange", "30m"))),
        "continent": str(data.get("continent") or "").strip() or None,
        "country": str(data.get("country") or "").strip() or None,
        "callsign": str(data.get("callsign") or "").strip() or None,
        "talkgroups": talkgroups,
        "rows": min(max(rows, 1), 100),
    }


def _live_qso_matches(row: dict[str, Any], subscription: dict[str, Any]) -> bool:
    start_at = row.get("start_at")
    if start_at is None:
        return False
    if start_at.tzinfo is None:
        start_at = start_at.replace(tzinfo=UTC)
    if start_at < start_time(subscription["time_range"]):
        return False
    if subscription["continent"] and row.get("continent") != subscription["continent"]:
        return False
    if subscription["country"] and row.get("country") != subscription["country"]:
        return False
    if subscription["talkgroups"] and row.get("destination_id") not in subscription["talkgroups"]:
        return False
    callsign = subscription["callsign"]
    if callsign and callsign.casefold() not in str(row.get("source_call") or "").casefold():
        return False
    return int(row.get("duration_ms") or 0) >= round(
        settings.kerchunk_threshold_seconds * 1000
    )


async def _next_qso_notification(connection: psycopg.AsyncConnection[Any]) -> str:
    async for notification in connection.notifies():
        return notification.payload
    raise RuntimeError("QSO notification stream ended")


async def _send_live_snapshot(
    websocket: WebSocket,
    subscription: dict[str, Any],
) -> None:
    rows = get_store().list_qsos(
        subscription["rows"],
        0,
        subscription["callsign"],
        list(subscription["talkgroups"]),
        start_time(subscription["time_range"]),
        subscription["continent"],
        subscription["country"],
        settings.kerchunk_threshold_seconds,
    )
    await websocket.send_json(
        {
            "type": "snapshot",
            "rows": jsonable_encoder([_public_qso(row) for row in rows]),
        }
    )


@app.websocket("/user/live-qsos/ws")
async def live_qsos_websocket(websocket: WebSocket) -> None:
    if _current_user(websocket) is None:
        await websocket.close(code=4401, reason="authentication required")
        return

    await websocket.accept()
    receive_task: asyncio.Task[Any] | None = None
    notify_task: asyncio.Task[Any] | None = None
    try:
        async with await psycopg.AsyncConnection.connect(settings.database_url) as connection:
            await connection.set_autocommit(True)
            await connection.execute(f"LISTEN {QSO_NOTIFY_CHANNEL}")
            receive_task = asyncio.create_task(websocket.receive_json())
            notify_task = asyncio.create_task(_next_qso_notification(connection))
            subscription: dict[str, Any] | None = None
            while True:
                done, _ = await asyncio.wait(
                    {receive_task, notify_task},
                    return_when=asyncio.FIRST_COMPLETED,
                )
                if receive_task in done:
                    message = receive_task.result()
                    receive_task = asyncio.create_task(websocket.receive_json())
                    if isinstance(message, dict) and message.get("type") in {
                        "subscribe",
                        "filters",
                    }:
                        subscription = _live_subscription(message)
                        await _send_live_snapshot(websocket, subscription)
                if notify_task in done:
                    session_id = notify_task.result()
                    notify_task = asyncio.create_task(_next_qso_notification(connection))
                    if subscription is None:
                        continue
                    row = get_store().get_live_qso(session_id)
                    if row is not None and _live_qso_matches(row, subscription):
                        await websocket.send_json(
                            {
                                "type": "qso",
                                "qso": jsonable_encoder(_public_qso(row)),
                            }
                        )
    except WebSocketDisconnect:
        return
    except (OSError, psycopg.Error, RuntimeError):
        with suppress(Exception):
            await websocket.close(code=1011, reason="live stream unavailable")
    finally:
        for task in (receive_task, notify_task):
            if task is not None:
                task.cancel()
        for task in (receive_task, notify_task):
            if task is not None:
                with suppress(asyncio.CancelledError, WebSocketDisconnect):
                    await task


@app.get("/user/live-qsos/data")
def live_qsos_data(
    request: Request,
    timeRange: str = "30m",
    continent: str | None = None,
    country: str | None = None,
    talkgroup: list[int] | None = Query(default=None),
    callsign: str | None = None,
    rows: int = Query(default=25, ge=1, le=100),
) -> Any:
    if _current_user(request) is None:
        return JSONResponse({"error": "authentication required"}, status_code=401)
    selected_range = _live_time_range(timeRange)
    selected_continent = continent or None
    selected_country = country or None
    result = get_store().list_qsos(
        rows,
        0,
        callsign,
        talkgroup,
        start_time(selected_range),
        selected_continent,
        selected_country,
        settings.kerchunk_threshold_seconds,
    )
    return [_public_qso(row) for row in result]


@app.get("/user/live-qsos", response_class=HTMLResponse)
def user_live_qsos(
    request: Request,
    timeRange: str = "30m",
    continent: str | None = None,
    country: str | None = None,
    talkgroup: list[int] | None = Query(default=None),
    callsign: str | None = None,
    rows: int = Query(default=25, ge=1, le=100),
) -> Response:
    locale = request_locale(request)
    user = _current_user(request)
    if user is None:
        return RedirectResponse("/user/login", status_code=303)

    selected_range = _live_time_range(timeRange)
    selected_continent = continent or None
    selected_country = country or None
    selected_rows = min(max(rows, 1), 100)
    selected_talkgroups = {str(value) for value in talkgroup or []}
    query_started = perf_counter()
    active_talkgroups = (
        get_store().active_talkgroups(
            start_time(selected_range), selected_continent, selected_country
        )
        if selected_continent and selected_country
        else []
    )
    live_rows = get_store().list_qsos(
        selected_rows,
        0,
        callsign,
        talkgroup,
        start_time(selected_range),
        selected_continent,
        selected_country,
        settings.kerchunk_threshold_seconds,
    )
    query_seconds = perf_counter() - query_started
    continents = get_store().continents()
    continent_options = "".join(
        f'<option value="{_escape(value)}"{" selected" if value == selected_continent else ""}>'
        f'{_escape(catalog(locale).get("metadata", {}).get("continents", {}).get(value, value))}</option>'
        for value in continents
    )
    country_labels = catalog(locale).get("metadata", {}).get("countries", {})
    country_options = "".join(
        f'<option value="{_escape(row["value"])}"{" selected" if row["value"] == selected_country else ""}>'
        f'{_escape(country_labels.get(row["value"], row["label"]))}</option>'
        for row in get_store().countries(selected_continent)
    )
    talkgroup_options = "".join(
        f'<option value="{_escape(row["value"])}"'
        f'{" selected" if str(row["value"]) in selected_talkgroups else ""}>'
        f'{_escape(row["label"])} ({_escape(row["count"])} { _escape(translate(locale, "home.qsos"))})</option>'
        for row in active_talkgroups
    )
    live_texts = json.dumps(
        {
            "noData": translate(locale, "live.noData"),
            "loadError": translate(locale, "live.loadError"),
            "updated": translate(locale, "live.updated"),
            "connecting": translate(locale, "live.connecting"),
            "connected": translate(locale, "live.connected"),
            "reconnecting": translate(locale, "live.reconnecting"),
            "disconnected": translate(locale, "live.disconnected"),
            "allCountries": translate(locale, "home.allCountries"),
            "justNow": translate(locale, "live.justNow"),
            "relativePrefix": translate(locale, "live.relativePrefix"),
            "relativeSuffix": translate(locale, "live.relativeSuffix"),
            "secondsUnit": translate(locale, "live.secondsUnit"),
            "minutesUnit": translate(locale, "live.minutesUnit"),
            "hoursUnit": translate(locale, "live.hoursUnit"),
            "daysUnit": translate(locale, "live.daysUnit"),
            "countries": catalog(locale).get("metadata", {}).get("countries", {}),
        }
    )
    content = f"""
<style>
.live-controls{{display:flex;gap:12px;flex-wrap:wrap;align-items:flex-start}}
.live-controls label{{display:block;margin:0 0 5px;font-size:12px;font-weight:800;color:#6b7280;text-transform:uppercase;letter-spacing:.04em}}
.live-controls select,.live-controls input,.live-controls button{{height:40px;padding:0 10px;border:2px solid #e5e7eb;border-radius:8px;background:#fff;font:inherit}}
.live-controls input{{min-width:185px}}.live-controls select[multiple]{{height:100px;min-width:240px}}
.live-controls button{{margin-top:21px;border:0;background:linear-gradient(135deg,#667eea,#764ba2);color:#fff;font-weight:700;cursor:pointer}}
.live-toggle{{height:40px;margin-top:21px;display:flex;align-items:center;gap:8px;color:#6b7280;font-size:13px;font-weight:600}}.live-toggle input{{width:auto;height:auto;margin:0;accent-color:#667eea}}
.live-table-wrap{{max-height:62vh;overflow:auto;border:1px solid #e8eaf0;border-radius:9px}}
.live-table{{min-width:760px;width:100%;border-collapse:collapse}}
.live-table th,.live-table td{{padding:8px 10px;border-bottom:1px solid #e8eaf0;text-align:left;font-size:13px;white-space:nowrap}}
.live-table th{{position:sticky;top:0;background:linear-gradient(135deg,#667eea,#764ba2);color:#fff;font-size:12px;text-transform:uppercase;letter-spacing:.04em}}
.live-table tbody tr:hover{{background:#faf9ff}}.live-primary{{font-weight:750;color:#5b5bd6}}.live-muted{{color:#6b7280;font-size:12px}}.live-duration{{font-family:ui-monospace,SFMono-Regular,Menlo,monospace;color:#6b7280}}
.live-status{{display:flex;justify-content:space-between;gap:12px;align-items:center;margin:12px 0 0;color:#6b7280;font-size:12px}}.live-status strong{{color:#15803d}}
@media(max-width:700px){{.live-controls{{flex-direction:column}}.live-controls>div,.live-controls input,.live-controls select,.live-controls button{{width:100%}}.live-controls button{{margin-top:8px}}}}
</style>
<section class="card"><div class="nav"><div><h1 style="margin:0;text-align:left">{_escape(translate(locale, "live.title"))}</h1><p class="muted">{_escape(translate(locale, "live.subtitle"))} {_escape(user["callsign"])}</p></div><div><a class="button secondary" href="/user/reports">{_escape(translate(locale, "home.reports"))}</a> <form class="inline" method="post" action="/user/logout"><button class="button secondary" type="submit">{_escape(translate(locale, "user.logout"))}</button></form></div></div>
<form id="liveForm" class="live-controls" method="get" action="/user/live-qsos">
<div><label for="liveContinent">{_escape(translate(locale, "home.continent"))}</label><select id="liveContinent" name="continent"><option value="">{_escape(translate(locale, "home.allContinents"))}</option>{continent_options}</select></div>
<div id="liveCountryControl"{"" if selected_continent else " style=\"display:none\""}><label for="liveCountry">{_escape(translate(locale, "home.country"))}</label><select id="liveCountry" name="country"><option value="">{_escape(translate(locale, "home.allCountries"))}</option>{country_options}</select></div>
<div><label for="liveCallsign">{_escape(translate(locale, "home.callsignFilter"))}</label><input id="liveCallsign" name="callsign" value="{_escape(callsign or "")}" placeholder="{_escape(translate(locale, "home.callsignPlaceholder"))}"></div>
<div id="liveTalkgroupControl"{"" if active_talkgroups else " style=\"display:none\""}><label for="liveTalkgroups">{_escape(translate(locale, "home.talkgroupFilter"))}</label><select id="liveTalkgroups" name="talkgroup" multiple>{talkgroup_options}</select></div>
<div><label for="liveRows">{_escape(translate(locale, "home.rows"))}</label><select id="liveRows" name="rows">{"".join(f'<option{" selected" if value == selected_rows else ""}>{value}</option>' for value in (10, 15, 25, 40, 50, 100))}</select></div>
</form></section>
<section class="card"><div class="live-table-wrap"><table class="live-table"><thead><tr><th>{_escape(translate(locale, "live.howLongAgo"))}</th><th>{_escape(translate(locale, "home.source"))}</th><th>{_escape(translate(locale, "home.talkgroup"))}</th><th>{_escape(translate(locale, "home.slot"))}</th><th>{_escape(translate(locale, "home.duration"))}</th></tr></thead><tbody id="liveQsoRows">{_live_qso_rows(live_rows, locale)}</tbody></table></div><div class="live-status"><span id="liveMode"><strong>{_escape(translate(locale, "live.connecting"))}</strong></span><span id="liveUpdated">—</span></div></section>
<script>
const liveTimeRange={json.dumps(selected_range)};
const liveRangeSeconds={TIME_RANGES[selected_range]};
const liveLocale={json.dumps(locale)};
const liveTexts={live_texts};
const liveForm=document.getElementById('liveForm'), liveRows=document.getElementById('liveQsoRows'), liveMode=document.getElementById('liveMode'), liveUpdated=document.getElementById('liveUpdated'), liveRowsSelect=document.getElementById('liveRows');
let liveSocket=null, liveReconnectTimer=null, liveCallsignTimer=null;
const liveEntries=new Map();
const liveEscape=value=>String(value??'').replace(/[&<>\"']/g,character=>({{'&':'&amp;','<':'&lt;','>':'&gt;','\"':'&quot;',"'":'&#39;'}}[character]));
const relativeTime=value=>{{const elapsed=Math.max(0,Math.floor((Date.now()-new Date(value).getTime())/1000));if(elapsed<5)return liveTexts.justNow;let amount,unit;if(elapsed<60){{amount=elapsed;unit='secondsUnit';}}else if(elapsed<3600){{amount=Math.floor(elapsed/60);unit='minutesUnit';}}else if(elapsed<86400){{amount=Math.floor(elapsed/3600);unit='hoursUnit';}}else{{amount=Math.floor(elapsed/86400);unit='daysUnit';}}return `${{liveTexts.relativePrefix}}${{amount}} ${{liveTexts[unit]}}${{liveTexts.relativeSuffix}}`;}};
const liveDuration=value=>{{const seconds=Math.round(Number(value||0)),hours=Math.floor(seconds/3600),minutes=Math.floor(seconds%3600/60),rest=seconds%60;return hours?`${{hours}}:${{String(minutes).padStart(2,'0')}}:${{String(rest).padStart(2,'0')}}`:minutes?`${{minutes}}:${{String(rest).padStart(2,'0')}}`:`${{rest}} sec`;}};
const selectedTalkgroups=()=>Array.from(document.getElementById('liveTalkgroups')?.selectedOptions||[]).map(option=>Number(option.value)).filter(Number.isFinite);
const subscription=()=>({{type:'subscribe',timeRange:liveTimeRange,continent:document.getElementById('liveContinent').value||null,country:document.getElementById('liveCountry').value||null,callsign:document.getElementById('liveCallsign').value.trim()||null,talkgroups:selectedTalkgroups(),rows:Number(liveRowsSelect.value)}});
function renderLive(){{const cutoff=Date.now()-liveRangeSeconds*1000;const rows=Array.from(liveEntries.values()).filter(row=>new Date(row.start).getTime()>=cutoff).sort((a,b)=>new Date(b.start)-new Date(a.start)).slice(0,Number(liveRowsSelect.value));liveRows.innerHTML=rows.length?rows.map(row=>`<tr><td>${{liveEscape(relativeTime(row.start))}}</td><td><span class="live-primary">${{liveEscape(row.sourceCall||row.sourceId||'—')}}</span> <span class="live-muted">${{liveEscape(row.sourceName||'')}}</span></td><td><span class="live-primary">${{liveEscape(row.destinationName||'—')}}</span> <span class="live-muted">(${{liveEscape(row.destinationId||'—')}})</span></td><td>${{liveEscape(row.slot||'—')}}</td><td class="live-duration">${{liveEscape(liveDuration(row.duration))}}</td></tr>`).join(''):`<tr><td colspan="5" class="muted">${{liveEscape(liveTexts.noData)}}</td></tr>`;}}
function setLiveMode(text){{liveMode.innerHTML=`<strong>${{liveEscape(text)}}</strong>`;}}
function sendSubscription(){{if(liveSocket?.readyState===WebSocket.OPEN)liveSocket.send(JSON.stringify(subscription()));}}
function handleLiveMessage(message){{if(message.type==='snapshot'){{liveEntries.clear();(message.rows||[]).forEach(row=>liveEntries.set(row.sessionId,row));renderLive();}}else if(message.type==='qso'&&message.qso){{liveEntries.set(message.qso.sessionId,message.qso);renderLive();}}liveUpdated.textContent=`${{liveTexts.updated}} ${{new Date().toLocaleTimeString(liveLocale)}}`;}}
function connectLive(){{if(liveSocket&&liveSocket.readyState<=WebSocket.OPEN)return;setLiveMode(liveTexts.connecting);const protocol=location.protocol==='https:'?'wss':'ws';liveSocket=new WebSocket(`${{protocol}}://${{location.host}}/user/live-qsos/ws`);liveSocket.addEventListener('open',()=>{{setLiveMode(liveTexts.connected);sendSubscription();}});liveSocket.addEventListener('message',event=>{{try{{handleLiveMessage(JSON.parse(event.data));}}catch(_){{}}}});liveSocket.addEventListener('error',()=>setLiveMode(liveTexts.disconnected));liveSocket.addEventListener('close',event=>{{if(event.code===4401){{window.location='/user/login';return;}}setLiveMode(liveTexts.reconnecting);clearTimeout(liveReconnectTimer);liveReconnectTimer=setTimeout(connectLive,2000);}});}}
async function loadCountries(){{const continent=document.getElementById('liveContinent').value;const country=document.getElementById('liveCountry');const countryControl=document.getElementById('liveCountryControl');if(!continent){{countryControl.style.display='none';country.innerHTML=`<option value="">${{liveEscape(liveTexts.allCountries||'All countries')}}</option>`;document.getElementById('liveTalkgroupControl').style.display='none';document.getElementById('liveTalkgroups').innerHTML='';sendSubscription();return;}}const previous=country.value;const response=await fetch('/public/countries?continent='+encodeURIComponent(continent),{{cache:'no-store'}});const countries=await response.json();country.innerHTML=`<option value="">${{liveEscape(liveTexts.allCountries||'All countries')}}</option>`+countries.map(item=>`<option value="${{liveEscape(item.value)}}">${{liveEscape(liveTexts.countries[item.value]||item.label)}}</option>`).join('');countryControl.style.display=countries.length?'block':'none';if(countries.some(item=>item.value===previous))country.value=previous;await loadTalkgroups();sendSubscription();}}
async function loadTalkgroups(){{const continent=document.getElementById('liveContinent').value,country=document.getElementById('liveCountry').value,control=document.getElementById('liveTalkgroupControl'),select=document.getElementById('liveTalkgroups');if(!continent||!country){{control.style.display='none';select.innerHTML='';return;}}const wanted=new Set(selectedTalkgroups().map(String));const params=new URLSearchParams({{timeRange:liveTimeRange,continent,country}});const response=await fetch('/user/talkgroups?'+params,{{cache:'no-store'}});if(!response.ok){{control.style.display='none';return;}}const groups=await response.json();select.innerHTML=groups.map(item=>`<option value="${{liveEscape(item.value)}}"${{wanted.has(String(item.value))?' selected':''}}>${{liveEscape(item.label)}} (${{liveEscape(item.count)}})</option>`).join('');control.style.display=groups.length?'block':'none';}}
document.getElementById('liveContinent').addEventListener('change',()=>loadCountries().catch(()=>setLiveMode(liveTexts.disconnected)));document.getElementById('liveCountry').addEventListener('change',()=>loadTalkgroups().then(sendSubscription).catch(()=>setLiveMode(liveTexts.disconnected)));document.getElementById('liveTalkgroups').addEventListener('change',sendSubscription);liveRowsSelect.addEventListener('change',()=>{{renderLive();sendSubscription();}});document.getElementById('liveCallsign').addEventListener('input',()=>{{clearTimeout(liveCallsignTimer);liveCallsignTimer=setTimeout(sendSubscription,300);}});liveForm.addEventListener('submit',event=>event.preventDefault());setInterval(renderLive,5000);renderLive();connectLive();
</script>
"""
    return _account_page_with_metrics(
        translate(locale, "live.title"),
        content,
        locale,
        records_retrieved=len(live_rows),
        query_seconds=query_seconds,
    )


def _report_time_range(value: str | None) -> str:
    return value if value in TIME_RANGES else "1M"


def _report_query(
    time_range: str,
    continent: str | None,
    country: str | None,
    talkgroups: list[int] | None,
    callsign: str | None,
) -> str:
    values: list[tuple[str, str]] = [("timeRange", time_range)]
    if callsign:
        values.append(("callsign", callsign))
    if continent:
        values.append(("continent", continent))
    if country:
        values.append(("country", country))
    for talkgroup in talkgroups or []:
        values.append(("talkgroup", str(talkgroup)))
    return urlencode(values)


def _report_duration(seconds: float | int) -> str:
    seconds = round(float(seconds or 0))
    hours, remainder = divmod(seconds, 3600)
    minutes, seconds = divmod(remainder, 60)
    if hours:
        return f"{hours}:{minutes:02d}:{seconds:02d}"
    if minutes:
        return f"{minutes}:{seconds:02d}"
    return f"{seconds} s"


def _report_histogram_label(value: Any, bucket_seconds: int) -> str:
    """Return a compact label that fits beneath a report histogram bar."""
    parsed = value
    if isinstance(value, str):
        try:
            parsed = datetime.fromisoformat(value.replace("Z", "+00:00"))
        except ValueError:
            parsed = None
    if isinstance(parsed, datetime):
        return parsed.strftime("%d/%m" if bucket_seconds >= 24 * 60 * 60 else "%H:%M")
    return str(value or "")[:8]


def _report_callsign_chart_rows(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    result = []
    for row in rows:
        callsign = str(row.get("callsign") or "—")
        name = str(row.get("source_name") or "").strip()
        result.append({**row, "report_label": f"{callsign} · {name}" if name else callsign})
    return result


REPORT_ENTRY_LIMIT = 50


def _limit_report_entries(report: dict[str, Any]) -> dict[str, Any]:
    """Keep report tables and exports bounded while retaining the full histogram."""
    limited = dict(report)
    limited["daily"] = list(report.get("daily", []))[:REPORT_ENTRY_LIMIT]
    limited["talkgroups"] = sorted(
        report.get("talkgroups", []),
        key=lambda row: (-int(row.get("qso_count") or 0), str(row.get("name") or "")),
    )[:REPORT_ENTRY_LIMIT]
    limited["callsigns"] = sorted(
        report.get("callsigns", []),
        key=lambda row: (-int(row.get("qso_count") or 0), str(row.get("callsign") or "")),
    )[:REPORT_ENTRY_LIMIT]
    return limited


def _report_bar_rows(rows: list[dict[str, Any]], label: str, value: str, formatter: Any, empty_label: str = "No data for this filter.") -> str:
    if not rows:
        return f'<p class="muted">{_escape(empty_label)}</p>'
    maximum = max(float(row.get(value) or 0) for row in rows) or 1
    result = []
    for row in rows:
        amount = float(row.get(value) or 0)
        width = max(2, amount / maximum * 100)
        result.append(
            f'<div class="report-bar"><span>{_escape(row.get(label, "—"))}</span>'
            f'<div><i style="width:{width:.1f}%"></i></div><strong>{_escape(formatter(amount))}</strong></div>'
        )
    return "".join(result)


def _report_histogram(rows: list[dict[str, Any]], bucket_seconds: int, locale: str, empty_label: str) -> str:
    if not rows or not any(int(row.get("qso_count") or 0) for row in rows):
        return f'<p class="muted">{_escape(empty_label)}</p>'
    maximum = max(int(row.get("qso_count") or 0) for row in rows) or 1
    columns = []
    label_step = max(1, len(rows) // 10)
    for index, row in enumerate(rows):
        bucket = row.get("bucket")
        label = _report_histogram_label(bucket, bucket_seconds)
        count = int(row.get("qso_count") or 0)
        height = max(2, count / maximum * 100) if count else 0
        visible_label = label if index % label_step == 0 or index == len(rows) - 1 else ""
        columns.append(
            f'<div class="histogram-column" title="{_escape(label)}: {_escape(count)} '
            f'{_escape(translate(locale, "reports.qsos"))}">'
            f'<span class="histogram-value">{_escape(count) if count else ""}</span>'
            f'<i style="height:{height:.1f}%"></i>'
            f'<small>{_escape(visible_label)}</small></div>'
        )
    return f'<div class="histogram">{"".join(columns)}</div>'


def _report_page(
    request: Request,
    user: dict[str, Any],
    time_range: str,
    continent: str | None,
    country: str | None,
    talkgroups: list[int] | None,
    callsign: str | None,
) -> HTMLResponse:
    locale = request_locale(request)
    query_started = perf_counter()
    report = get_store().user_report(
        callsign,
        start_time(time_range),
        continent,
        country,
        talkgroups,
        histogram_bucket_seconds(time_range),
    )
    report = _limit_report_entries(report)
    translations = catalog(locale)
    metadata = translations.get("metadata", {})
    country_labels = metadata.get("countries", {})
    continents = get_store().continents()
    countries = get_store().countries(continent)
    query_seconds = perf_counter() - query_started
    selected_talkgroups = {str(value) for value in talkgroups or []}
    range_keys = {
        "5m": "home.last5m", "15m": "home.last15m", "30m": "home.last30m",
        "1h": "home.last1h", "2h": "home.last2h", "6h": "home.last6h",
        "12h": "home.last12h", "24h": "home.last24h", "2d": "home.last2d",
        "5d": "home.last5d", "1w": "home.last1w", "2w": "home.last2w",
        "1M": "home.last1M", "2M": "home.last2M", "3M": "home.last3M",
    }
    range_options = "".join(
        f'<option value="{key}"{" selected" if key == time_range else ""}>'
        f'{_escape(translate(locale, label_key, key))}</option>'
        for key, label_key in range_keys.items()
    )
    continent_options = "".join(
        f'<option value="{_escape(value)}"{" selected" if value == continent else ""}>'
        f'{_escape(metadata.get("continents", {}).get(value, value))}</option>'
        for value in continents
    )
    country_options = "".join(
        f'<option value="{_escape(row["value"])}"{" selected" if row["value"] == country else ""}>'
        f'{_escape(country_labels.get(row["value"], row["label"]))}</option>'
        for row in countries
    )
    report["talkgroups"] = sorted(
        report["talkgroups"],
        key=lambda row: (-int(row.get("qso_count") or 0), str(row.get("name") or ""), int(row.get("talkgroup_id") or 0)),
    )
    talkgroup_options = "".join(
        f'<option value="{_escape(row["talkgroup_id"])}"'
        f'{" selected" if str(row["talkgroup_id"]) in selected_talkgroups else ""}>'
        f'{_escape(row["name"])} ({_escape(row["qso_count"])} QSOs)</option>'
        for row in report["talkgroups"]
    )
    query = _report_query(time_range, continent, country, talkgroups, callsign)
    report_scope = callsign.strip() if callsign and callsign.strip() else translate(locale, "reports.allCallsigns")
    no_data = translate(locale, "reports.noData")
    summary = report["summary"]
    daily_rows = "".join(
        f'<tr><td>{_escape(row["day"])}</td><td>{_escape(row["qso_count"])}</td>'
        f'<td>{_escape(_report_duration(row["duration_seconds"]))}</td></tr>'
        for row in report["daily"]
    ) or f'<tr><td colspan="3" class="muted">{_escape(no_data)}</td></tr>'
    talkgroup_rows = "".join(
        f'<tr><td>{_escape(row["name"])}</td><td>{_escape(row["talkgroup_id"])}</td>'
        f'<td>{_escape(row["qso_count"])}</td><td>{_escape(_report_duration(row["duration_seconds"]))}</td>'
        f'<td>{_escape(_format_datetime(row["last_seen_at"]))}</td></tr>'
        for row in report["talkgroups"]
    ) or f'<tr><td colspan="5" class="muted">{_escape(no_data)}</td></tr>'
    callsign_chart_rows = _report_callsign_chart_rows(report["callsigns"])
    callsign_rows = "".join(
        f'<tr><td>{_escape(row["callsign"])}</td>'
        f'<td>{_escape(row.get("source_name") or "—")}</td>'
        f'<td>{_escape(row.get("countries") or "—")}</td>'
        f'<td>{_escape(row["qso_count"])}</td>'
        f'<td>{_escape(_report_duration(row["duration_seconds"]))}</td>'
        f'<td>{_escape(row["unique_talkgroups"])}</td></tr>'
        for row in report["callsigns"]
    ) or f'<tr><td colspan="6" class="muted">{_escape(no_data)}</td></tr>'
    content = f"""
<style>.report-controls{{display:flex;gap:12px;flex-wrap:wrap;align-items:flex-start}}.report-controls label{{display:block;font-size:12px;font-weight:700;color:#6b7280;margin-bottom:5px}}.report-controls select,.report-controls input,.report-controls button{{height:40px;padding:0 10px;border:2px solid #e5e7eb;border-radius:8px;background:#fff;font:inherit}}.report-controls input{{min-width:180px}}.report-controls select[multiple]{{height:100px;min-width:220px}}.report-controls button{{margin-top:21px;border:0;background:linear-gradient(135deg,#667eea,#764ba2);color:#fff;font-weight:700;cursor:pointer}}.report-actions{{display:flex;gap:9px;flex-wrap:wrap;margin-top:16px}}.report-bar{{display:grid;grid-template-columns:220px 1fr 75px;gap:8px;align-items:center;margin:8px 0;font-size:12px}}.report-bar span{{overflow:hidden;text-overflow:ellipsis;white-space:nowrap}}.report-bar div{{height:18px;background:#f0f1f6;border-radius:5px;overflow:hidden}}.report-bar i{{display:block;height:100%;background:linear-gradient(90deg,#667eea,#764ba2);border-radius:5px}}.report-bar strong{{text-align:right;color:#6b7280;white-space:nowrap}}.report-callsign-table th,.report-callsign-table td{{white-space:nowrap}}.histogram{{height:240px;display:flex;align-items:end;gap:3px;padding:18px 4px 28px;border-bottom:1px solid #e5e7eb;overflow:hidden}}.histogram-column{{position:relative;display:flex;flex:1;min-width:8px;height:100%;align-items:end;justify-content:end}}.histogram-column i{{display:block;width:100%;min-height:0;background:linear-gradient(180deg,#764ba2,#667eea);border-radius:4px 4px 0 0}}.histogram-column small{{position:absolute;bottom:-24px;left:50%;transform:translateX(-50%);width:56px;max-width:56px;overflow:hidden;text-overflow:ellipsis;font-size:9px;color:#6b7280;text-align:center;white-space:nowrap}}.histogram-value{{position:absolute;top:-16px;font-size:10px;color:#6b7280}}@media(max-width:700px){{.report-bar{{grid-template-columns:160px 1fr 55px}}.histogram{{gap:2px}}.histogram-value{{display:none}}}}</style>
<section class="card"><div class="nav"><div><h1 style="margin:0;text-align:left">{_escape(translate(locale, "reports.title"))}</h1><p class="muted">{_escape(translate(locale, "reports.forUser"))}: {_escape(user['callsign'])} · {_escape(translate(locale, "reports.scope"))}: {_escape(report_scope)}</p></div><div><a class="button secondary" href="/user/live-qsos">{_escape(translate(locale, "live.title"))}</a> <form class="inline" method="post" action="/user/logout"><button class="button secondary" type="submit">{_escape(translate(locale, "user.logout"))}</button></form></div></div>
<form id="reportForm" class="report-controls" method="get" action="/user/reports"><div><label>{_escape(translate(locale, "reports.callsign"))}</label><input name="callsign" value="{_escape(callsign or '')}" placeholder="{_escape(translate(locale, "home.callsignPlaceholder"))}"></div><div><label>{_escape(translate(locale, "reports.timeRange"))}</label><select id="reportTimeRange" name="timeRange">{range_options}</select></div><div><label>{_escape(translate(locale, "home.continent"))}</label><select id="reportContinent" name="continent"><option value="">{_escape(translate(locale, "home.allContinents"))}</option>{continent_options}</select></div><div><label>{_escape(translate(locale, "home.country"))}</label><select id="reportCountry" name="country"><option value="">{_escape(translate(locale, "home.allCountries"))}</option>{country_options}</select></div><div><label>{_escape(translate(locale, "reports.talkgroups"))}</label><select id="reportTalkgroups" name="talkgroup" multiple>{talkgroup_options}</select></div><button type="submit">{_escape(translate(locale, "reports.generate"))}</button></form>
<div class="report-actions"><a class="button secondary" href="/user/reports/export.csv?{query}">{_escape(translate(locale, "reports.csv"))}</a><a class="button secondary" href="/user/reports/export.xlsx?{query}">{_escape(translate(locale, "reports.excel"))}</a><a class="button" href="/user/reports/export.pdf?{query}">{_escape(translate(locale, "reports.pdf"))}</a></div></section>
<section class="card"><h2>{_escape(translate(locale, "reports.summary"))}</h2><div class="stats"><div class="stat"><small>{_escape(translate(locale, "reports.qsos"))}</small><strong>{_escape(summary['qso_count'])}</strong></div><div class="stat"><small>{_escape(translate(locale, "reports.talkTime"))}</small><strong>{_escape(_report_duration(summary['duration_seconds']))}</strong></div><div class="stat"><small>{_escape(translate(locale, "reports.uniqueTalkgroups"))}</small><strong>{_escape(summary['unique_talkgroups'])}</strong></div><div class="stat"><small>{_escape(translate(locale, "reports.activeDays"))}</small><strong>{_escape(summary['active_days'])}</strong></div></div></section>
<section class="card"><h2>{_escape(translate(locale, "reports.dailyActivity"))}</h2>{_report_histogram(report['histogram'], histogram_bucket_seconds(time_range), locale, no_data)}</section>
<section class="card"><h2>{_escape(translate(locale, "reports.talkgroupActivity"))}</h2>{_report_bar_rows(report['talkgroups'], 'name', 'qso_count', lambda value: f'{int(value)} {translate(locale, "reports.qsos")}', no_data)}<div class="table-wrap"><table><thead><tr><th>{_escape(translate(locale, "home.talkgroup"))}</th><th>{_escape(translate(locale, "user.id"))}</th><th>{_escape(translate(locale, "reports.qsos"))}</th><th>{_escape(translate(locale, "reports.talkTime"))}</th><th>{_escape(translate(locale, "home.lastHeard"))}</th></tr></thead><tbody>{talkgroup_rows}</tbody></table></div></section>
<section class="card"><h2>{_escape(translate(locale, "reports.callsignActivity"))}</h2>{_report_bar_rows(callsign_chart_rows, 'report_label', 'qso_count', lambda value: f'{int(value)} {translate(locale, "reports.qsos")}', no_data)}<div class="table-wrap"><table class="report-callsign-table"><thead><tr><th>{_escape(translate(locale, "home.callsignFilter"))}</th><th>{_escape(translate(locale, "user.name"))}</th><th>{_escape(translate(locale, "home.country"))}</th><th>{_escape(translate(locale, "reports.qsos"))}</th><th>{_escape(translate(locale, "reports.talkTime"))}</th><th>{_escape(translate(locale, "reports.uniqueTalkgroups"))}</th></tr></thead><tbody>{callsign_rows}</tbody></table></div></section>
<section class="card"><h2>{_escape(translate(locale, "reports.dailyTable"))}</h2><div class="table-wrap"><table><thead><tr><th>{_escape(translate(locale, "reports.date"))}</th><th>{_escape(translate(locale, "reports.qsos"))}</th><th>{_escape(translate(locale, "reports.talkTime"))}</th></tr></thead><tbody>{daily_rows}</tbody></table></div></section>
<script>document.getElementById('reportContinent')?.addEventListener('change',()=>document.getElementById('reportForm').submit());document.getElementById('reportCountry')?.addEventListener('change',()=>document.getElementById('reportForm').submit());document.getElementById('reportTimeRange')?.addEventListener('change',()=>document.getElementById('reportForm').submit());</script>
"""
    return _account_page_with_metrics(
        translate(locale, "reports.title"),
        content,
        locale,
        records_retrieved=report["summary"]["qso_count"],
        query_seconds=query_seconds,
    )


def _load_user_report(
    request: Request,
    time_range: str,
    continent: str | None,
    country: str | None,
    talkgroups: list[int] | None,
    callsign: str | None,
) -> tuple[dict[str, Any] | None, dict[str, Any] | None]:
    user = _current_user(request)
    if user is None:
        return None, None
    report = get_store().user_report(
        callsign,
        start_time(time_range),
        continent,
        country,
        talkgroups,
        histogram_bucket_seconds(time_range),
    )
    return user, _limit_report_entries(report)


def _report_csv(report: dict[str, Any]) -> bytes:
    report = _limit_report_entries(report)
    output = io.StringIO(newline="")
    writer = csv.writer(output)
    summary = report["summary"]
    writer.writerow(["Metric", "Value"])
    writer.writerow(["QSOs", summary["qso_count"]])
    writer.writerow(["Talk time seconds", summary["duration_seconds"]])
    writer.writerow(["Unique talkgroups", summary["unique_talkgroups"]])
    writer.writerow(["Active days", summary["active_days"]])
    writer.writerow([])
    writer.writerow(["Callsign", "Name", "Country", "QSOs", "Talk time seconds", "Unique talkgroups"])
    for row in report["callsigns"]:
        writer.writerow([
            row["callsign"], row.get("source_name") or "", row.get("countries") or "",
            row["qso_count"], row["duration_seconds"], row["unique_talkgroups"],
        ])
    return output.getvalue().encode("utf-8-sig")


def _report_excel(report: dict[str, Any]) -> bytes:
    from openpyxl import Workbook

    report = _limit_report_entries(report)
    workbook = Workbook()
    summary_sheet = workbook.active
    summary_sheet.title = "Summary"
    summary = report["summary"]
    summary_sheet.append(["Metric", "Value"])
    for label, value in (
        ("QSOs", summary["qso_count"]),
        ("Talk time seconds", summary["duration_seconds"]),
        ("Unique talkgroups", summary["unique_talkgroups"]),
        ("Active days", summary["active_days"]),
        ("First QSO", summary["first_qso_at"]),
        ("Last QSO", summary["last_qso_at"]),
    ):
        if isinstance(value, datetime) and value.tzinfo:
            value = value.replace(tzinfo=None)
        summary_sheet.append([label, value])
    daily_sheet = workbook.create_sheet("Daily activity")
    daily_sheet.append(["Date", "QSOs", "Talk time seconds"])
    for row in report["daily"]:
        daily_sheet.append([row["day"], row["qso_count"], row["duration_seconds"]])
    talkgroup_sheet = workbook.create_sheet("Talkgroups")
    talkgroup_sheet.append(["Talkgroup", "ID", "QSOs", "Talk time seconds", "Last seen"])
    for row in report["talkgroups"]:
        last_seen = row["last_seen_at"]
        if isinstance(last_seen, datetime) and last_seen.tzinfo:
            last_seen = last_seen.replace(tzinfo=None)
        talkgroup_sheet.append([row["name"], row["talkgroup_id"], row["qso_count"], row["duration_seconds"], last_seen])
    callsign_sheet = workbook.create_sheet("Callsigns")
    callsign_sheet.append(["Callsign", "Name", "Country", "QSOs", "Talk time seconds", "Unique talkgroups", "Last seen"])
    for row in report["callsigns"]:
        last_seen = row["last_seen_at"]
        if isinstance(last_seen, datetime) and last_seen.tzinfo:
            last_seen = last_seen.replace(tzinfo=None)
        callsign_sheet.append([
            row["callsign"], row.get("source_name") or "", row.get("countries") or "",
            row["qso_count"], row["duration_seconds"], row["unique_talkgroups"], last_seen,
        ])
    for sheet in workbook.worksheets:
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions
        for cell in sheet[1]:
            cell.font = cell.font.copy(bold=True)
    output = io.BytesIO()
    workbook.save(output)
    return output.getvalue()


def _pdf_histogram(
    reportlab_rows: list[dict[str, Any]],
    width: float,
    bucket_seconds: int = 86400,
) -> Any | None:
    from reportlab.lib.colors import HexColor
    from reportlab.graphics.shapes import Drawing, Line, Rect, String

    rows = reportlab_rows
    if not rows or not any(int(row.get("qso_count") or 0) for row in rows):
        return None
    height = 128
    left, bottom, chart_width, chart_height = 30, 28, width - 42, 86
    maximum = max(int(row.get("qso_count") or 0) for row in rows) or 1
    drawing = Drawing(width, height)
    drawing.add(Line(left, bottom, left + chart_width, bottom, strokeColor=HexColor("#9ca3af")))
    bar_width = chart_width / max(len(rows), 1)
    label_step = max(1, len(rows) // 8)
    for index, row in enumerate(rows):
        count = int(row.get("qso_count") or 0)
        x = left + index * bar_width + 0.5
        bar_height = count / maximum * chart_height if count else 0
        drawing.add(Rect(x, bottom, max(1, bar_width - 1), bar_height, fillColor=HexColor("#667eea"), strokeColor=None))
        if index % label_step == 0 or index == len(rows) - 1:
            label = _report_histogram_label(row.get("bucket"), bucket_seconds)
            drawing.add(String(x, 10, label, fontSize=6, fillColor="#6b7280"))
    return drawing


def _pdf_bar_chart(rows: list[dict[str, Any]], label: str, value: str, width: float) -> Any | None:
    from reportlab.lib.colors import HexColor
    from reportlab.graphics.shapes import Drawing, Rect, String

    rows = rows[:10]
    if not rows:
        return None
    left, row_height, bar_width = 115, 20, width - 150
    height = row_height * len(rows) + 6
    maximum = max(float(row.get(value) or 0) for row in rows) or 1
    drawing = Drawing(width, height)
    for index, row in enumerate(rows):
        y = height - (index + 1) * row_height + 4
        text = str(row.get(label, "—"))
        if len(text) > 19:
            text = text[:18] + "…"
        amount = float(row.get(value) or 0)
        drawing.add(String(0, y + 2, text, fontSize=7, fillColor=HexColor("#374151")))
        drawing.add(Rect(left, y, max(1, amount / maximum * bar_width), 10, fillColor=HexColor("#764ba2"), strokeColor=None))
        drawing.add(String(width - 30, y + 2, str(int(amount)), fontSize=7, fillColor=HexColor("#6b7280")))
    return drawing


def _report_pdf(
    report: dict[str, Any],
    callsign: str,
    locale: str,
    bucket_seconds: int = 86400,
) -> bytes:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.lib.units import mm
    from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

    report = _limit_report_entries(report)
    output = io.BytesIO()
    document = SimpleDocTemplate(output, pagesize=A4, rightMargin=14 * mm, leftMargin=14 * mm, topMargin=14 * mm, bottomMargin=14 * mm)
    styles = getSampleStyleSheet()
    story: list[Any] = [
        Paragraph(_escape(translate(locale, "reports.title")), styles["Title"]),
        Paragraph(f"{_escape(translate(locale, 'reports.forUser'))}: {_escape(callsign)}", styles["Normal"]),
        Spacer(1, 8),
    ]
    summary = report["summary"]
    story.append(Table(
        [[translate(locale, "reports.qsos"), summary["qso_count"], translate(locale, "reports.talkTime"), _report_duration(summary["duration_seconds"])],
         [translate(locale, "reports.uniqueTalkgroups"), summary["unique_talkgroups"], translate(locale, "reports.activeDays"), summary["active_days"]]],
        colWidths=[38 * mm, 32 * mm, 42 * mm, 45 * mm],
        style=TableStyle([("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#f1f3ff")), ("GRID", (0, 0), (-1, -1), 0.25, colors.lightgrey), ("FONTNAME", (0, 0), (-1, -1), "Helvetica"), ("VALIGN", (0, 0), (-1, -1), "MIDDLE")]),
    ))
    story.append(Spacer(1, 10))
    story.append(Paragraph(_escape(translate(locale, "reports.dailyActivity")), styles["Heading2"]))
    histogram = _pdf_histogram(report["histogram"], 170 * mm, bucket_seconds)
    if histogram is not None:
        story.append(histogram)
    story.append(Spacer(1, 8))
    story.append(Paragraph(_escape(translate(locale, "reports.talkgroupActivity")), styles["Heading2"]))
    talkgroup_chart = _pdf_bar_chart(report["talkgroups"], "name", "qso_count", 170 * mm)
    if talkgroup_chart is not None:
        story.append(talkgroup_chart)
        story.append(Spacer(1, 6))
    talkgroup_table = [[translate(locale, "home.talkgroup"), translate(locale, "user.id"), translate(locale, "reports.qsos"), translate(locale, "reports.talkTime")]] + [
        [str(row["name"]), str(row["talkgroup_id"]), str(row["qso_count"]), _report_duration(row["duration_seconds"])]
        for row in report["talkgroups"]
    ]
    if len(talkgroup_table) == 1:
        talkgroup_table.append([translate(locale, "reports.noData"), "", "", ""])
    story.append(Table(talkgroup_table, repeatRows=1, style=TableStyle([("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#667eea")), ("TEXTCOLOR", (0, 0), (-1, 0), colors.white), ("GRID", (0, 0), (-1, -1), 0.25, colors.lightgrey), ("FONTSIZE", (0, 0), (-1, -1), 8)])))
    story.append(Spacer(1, 10))
    story.append(Paragraph(_escape(translate(locale, "reports.callsignActivity")), styles["Heading2"]))
    callsign_chart = _pdf_bar_chart(
        _report_callsign_chart_rows(report["callsigns"]),
        "report_label",
        "qso_count",
        170 * mm,
    )
    if callsign_chart is not None:
        story.append(callsign_chart)
        story.append(Spacer(1, 6))
    callsign_table = [[translate(locale, "reports.callsign"), translate(locale, "user.name"), translate(locale, "home.country"), translate(locale, "reports.qsos"), translate(locale, "reports.talkTime"), translate(locale, "reports.talkgroups")]] + [
        [str(row["callsign"]), str(row.get("source_name") or "—"), str(row.get("countries") or "—"), str(row["qso_count"]), _report_duration(row["duration_seconds"]), str(row["unique_talkgroups"])]
        for row in report["callsigns"]
    ]
    if len(callsign_table) == 1:
        callsign_table.append([translate(locale, "reports.noData"), "", "", "", "", ""])
    story.append(Table(callsign_table, repeatRows=1, style=TableStyle([("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#667eea")), ("TEXTCOLOR", (0, 0), (-1, 0), colors.white), ("GRID", (0, 0), (-1, -1), 0.25, colors.lightgrey), ("FONTSIZE", (0, 0), (-1, -1), 8)])))
    story.append(Spacer(1, 10))
    story.append(Paragraph(_escape(translate(locale, "reports.dailyTable")), styles["Heading2"]))
    daily_table = [[translate(locale, "reports.date"), translate(locale, "reports.qsos"), translate(locale, "reports.talkTime")]] + [[str(row["day"]), str(row["qso_count"]), _report_duration(row["duration_seconds"])] for row in report["daily"]]
    if len(daily_table) == 1:
        daily_table.append([translate(locale, "reports.noData"), "", ""])
    story.append(Table(daily_table, repeatRows=1, style=TableStyle([("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#667eea")), ("TEXTCOLOR", (0, 0), (-1, 0), colors.white), ("GRID", (0, 0), (-1, -1), 0.25, colors.lightgrey), ("FONTSIZE", (0, 0), (-1, -1), 8)])))
    document.build(story)
    return output.getvalue()


@app.get("/user/reports", response_class=HTMLResponse)
def user_reports(
    request: Request,
    timeRange: str = "1M",
    continent: str | None = None,
    country: str | None = None,
    talkgroup: list[int] | None = Query(default=None),
    callsign: str | None = None,
) -> Response:
    user = _current_user(request)
    if user is None:
        return RedirectResponse("/user/login", status_code=303)
    return _report_page(request, user, _report_time_range(timeRange), continent, country, talkgroup, callsign)


def _report_export(
    request: Request,
    time_range: str,
    continent: str | None,
    country: str | None,
    talkgroups: list[int] | None,
    callsign: str | None,
    kind: str,
) -> Response:
    user, report = _load_user_report(request, _report_time_range(time_range), continent, country, talkgroups, callsign)
    if user is None or report is None:
        return JSONResponse({"error": "authentication required"}, status_code=401)
    if kind == "csv":
        return Response(_report_csv(report), media_type="text/csv; charset=utf-8", headers={"Content-Disposition": f'attachment; filename="{user["callsign"]}-report.csv"'})
    if kind == "xlsx":
        return Response(_report_excel(report), media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers={"Content-Disposition": f'attachment; filename="{user["callsign"]}-report.xlsx"'})
    scope = callsign.strip() if callsign and callsign.strip() else translate(request_locale(request), "reports.allCallsigns")
    return Response(_report_pdf(report, scope, request_locale(request), histogram_bucket_seconds(_report_time_range(time_range))), media_type="application/pdf", headers={"Content-Disposition": f'attachment; filename="{user["callsign"]}-report.pdf"'})


@app.get("/user/reports/export.csv")
def user_report_csv(request: Request, timeRange: str = "1M", continent: str | None = None, country: str | None = None, talkgroup: list[int] | None = Query(default=None), callsign: str | None = None) -> Response:
    return _report_export(request, timeRange, continent, country, talkgroup, callsign, "csv")


@app.get("/user/reports/export.xlsx")
def user_report_excel(request: Request, timeRange: str = "1M", continent: str | None = None, country: str | None = None, talkgroup: list[int] | None = Query(default=None), callsign: str | None = None) -> Response:
    return _report_export(request, timeRange, continent, country, talkgroup, callsign, "xlsx")


@app.get("/user/reports/export.pdf")
def user_report_pdf(request: Request, timeRange: str = "1M", continent: str | None = None, country: str | None = None, talkgroup: list[int] | None = Query(default=None), callsign: str | None = None) -> Response:
    return _report_export(request, timeRange, continent, country, talkgroup, callsign, "pdf")


@app.get("/user/api/stats")
def user_api_stats(request: Request) -> JSONResponse:
    user = _current_user(request)
    if user is None:
        return JSONResponse({"error": "authentication required"}, status_code=401)
    return JSONResponse(jsonable_encoder(get_store().user_statistics(user["callsign"])))


@app.post("/user/change-password")
async def user_change_password(request: Request) -> Response:
    locale = request_locale(request)
    user = _current_user(request)
    if user is None:
        return RedirectResponse("/user/login", status_code=303)
    fields = await _form_fields(request)
    if not verify_password(fields.get("current_password", ""), user["password_hash"]):
        return _account_page(translate(locale, "user.profile"), f'<section class="card"><p class="error">{_escape(translate(locale, "user.currentPasswordIncorrect"))}</p><a class="button" href="/user/profile">{_escape(translate(locale, "user.profile"))}</a></section>', locale)
    if len(fields.get("new_password", "")) < 8:
        return _account_page(translate(locale, "user.profile"), f'<section class="card"><p class="error">{_escape(translate(locale, "user.newPasswordMin"))}</p><a class="button" href="/user/profile">{_escape(translate(locale, "user.profile"))}</a></section>', locale)
    get_store().update_user_password(user["id"], hash_password(fields["new_password"]))
    return RedirectResponse("/user/profile", status_code=303)


@app.get("/admin/login", response_class=HTMLResponse)
def admin_login_page(request: Request) -> HTMLResponse:
    locale = request_locale(request)
    message = translate(locale, "admin.setPassword") if not settings.admin_password else ""
    warning = f'<p class="error">{message}</p>' if message else ""
    content = f'<section class="card form"><h1>{_escape(translate(locale, "admin.login"))}</h1>{warning}<form method="post" action="/admin/login"><label>{_escape(translate(locale, "admin.password"))}</label><input type="password" name="password" required><button class="button" type="submit">{_escape(translate(locale, "admin.open"))}</button></form></section>'
    return _account_page(translate(locale, "admin.login"), content, locale)


@app.post("/admin/login")
async def admin_login(request: Request) -> Response:
    locale = request_locale(request)
    fields = await _form_fields(request)
    if not settings.admin_password or not hmac.compare_digest(fields.get("password", ""), settings.admin_password):
        return _account_page(translate(locale, "admin.login"), f'<section class="card"><p class="error">{_escape(translate(locale, "admin.invalidPassword"))}</p><a class="button" href="/admin/login">{_escape(translate(locale, "admin.open"))}</a></section>', locale)
    response = _admin_redirect()
    response.set_cookie("admin_session", issue_admin_token(settings.admin_password), max_age=8 * 60 * 60, httponly=True, samesite="lax", secure=settings.cookie_secure)
    return response


@app.post("/admin/logout")
def admin_logout() -> RedirectResponse:
    response = RedirectResponse("/", status_code=303)
    response.delete_cookie("admin_session")
    return response


def _admin_user_row(user: dict[str, Any], locale: str = "en") -> str:
    active = bool(user["is_active"])
    status = translate(locale, "admin.active" if active else "admin.inactive")
    action = translate(locale, "admin.deactivate" if active else "admin.activate")
    active_value = "false" if active else "true"
    confirm_delete = _escape(json.dumps(translate(locale, "admin.confirmDelete")))
    return f"""
<tr><td>{_escape(user['callsign'])}</td><td>{_escape(user['name'])}</td><td>{_escape(user['email'])}</td>
<td>{status}</td><td>{_escape(user['qso_count'])}</td><td>{_escape(round(user['duration_seconds'], 1))} s</td>
<td><form class="inline" method="post" action="/admin/users/{user['id']}/status"><input type="hidden" name="active" value="{active_value}"><button class="button secondary" type="submit">{_escape(action)}</button></form>
<form class="inline" method="post" action="/admin/users/{user['id']}/delete" onsubmit="return window.confirm({confirm_delete})"><button class="button danger" type="submit">{_escape(translate(locale, "admin.delete"))}</button></form></td></tr>
"""


def _postgres_table_rows(overview: dict[str, Any], locale: str = "en") -> str:
    rows = overview.get("tables", [])
    return "".join(
        f"<tr><td>{_escape(row['table_name'])}</td>"
        f"<td>{_escape(row['estimated_rows'])}</td>"
        f"<td>{_escape(row['total_size'])}</td></tr>"
        for row in rows
    ) or f'<tr><td colspan="3" class="muted">{_escape(translate(locale, "admin.noTables"))}</td></tr>'


def _postgres_connection_rows(overview: dict[str, Any], locale: str = "en") -> str:
    rows = overview.get("connection_states", [])
    return "".join(
        f"<tr><td>{_escape(row['state'])}</td><td>{_escape(row['count'])}</td></tr>"
        for row in rows
    ) or f'<tr><td colspan="2" class="muted">{_escape(translate(locale, "admin.noConnections"))}</td></tr>'


ADMIN_RETENTION_MONTHS = (1, 2, 3, 6)


def _admin_retention_period(locale: str, months: int) -> str:
    unit = translate(
        locale,
        "adminMaintenance.month" if months == 1 else "adminMaintenance.months",
    )
    return f"{months} {unit}"


def _admin_maintenance_notice(request: Request, locale: str) -> str:
    notice = request.query_params.get("notice")
    if not notice:
        return ""
    try:
        if notice == "rebuild":
            message = translate(locale, "adminMaintenance.rebuildSuccess").format(
                count=int(request.query_params.get("count", "0")),
                raw=int(request.query_params.get("raw", "0")),
            )
        elif notice == "raw":
            months = int(request.query_params.get("months", "0"))
            message = translate(locale, "adminMaintenance.rawDeleteSuccess").format(
                count=int(request.query_params.get("count", "0")),
                period=_admin_retention_period(locale, months),
                qsos=int(request.query_params.get("qsos", "0")),
            )
        elif notice == "qso":
            months = int(request.query_params.get("months", "0"))
            message = translate(locale, "adminMaintenance.qsoDeleteSuccess").format(
                count=int(request.query_params.get("count", "0")),
                period=_admin_retention_period(locale, months),
            )
        else:
            return ""
    except (TypeError, ValueError):
        return ""
    return f'<p class="success"><strong>{_escape(translate(locale, "adminMaintenance.success"))}:</strong> {_escape(message)}</p>'


def _admin_retention_row(
    locale: str,
    kind: str,
    months: int,
    counts: dict[str, int],
) -> str:
    if kind == "raw-events":
        count = counts["raw_events"]
        detail = f'{count} raw events · {counts["dependent_qsos"]} {translate(locale, "adminMaintenance.dependentQsos")}'
        confirmation = translate(locale, "adminMaintenance.rawConfirm").format(
            raw=count,
            period=_admin_retention_period(locale, months),
            qsos=counts["dependent_qsos"],
        )
    else:
        count = counts["qsos"]
        detail = f'{count} QSOs'
        confirmation = translate(locale, "adminMaintenance.qsoConfirm").format(
            qsos=count,
            period=_admin_retention_period(locale, months),
        )
    confirm_json = _escape(json.dumps(confirmation))
    return f'''
<div class="nav" style="margin:10px 0;align-items:center"><div><strong>{_escape(translate(locale, "adminMaintenance.olderThan"))} {_escape(_admin_retention_period(locale, months))}</strong><br><span class="muted">{_escape(translate(locale, "adminMaintenance.eligible"))}: {_escape(detail)}</span></div>
<form method="post" action="/admin/maintenance/{kind}/{months}" onsubmit="return window.confirm({confirm_json})"><button class="button danger" type="submit">{_escape(translate(locale, "adminMaintenance.deleteButton"))}</button></form></div>'''


@app.get("/admin", response_class=HTMLResponse)
def admin_panel(request: Request) -> Response:
    locale = request_locale(request)
    if not _admin_allowed(request):
        return RedirectResponse("/admin/login", status_code=303)
    query_started = perf_counter()
    stats = get_store().admin_statistics()
    users = get_store().list_users()
    postgres = get_store().postgres_overview()
    maintenance = get_store().maintenance_overview()
    retention = {months: get_store().retention_counts(months) for months in ADMIN_RETENTION_MONTHS}
    query_seconds = perf_counter() - query_started
    rows = "".join(_admin_user_row(user, locale) for user in users) or f'<tr><td colspan="7" class="muted">{_escape(translate(locale, "admin.registeredUsers"))}</td></tr>'
    postgres_tables = _postgres_table_rows(postgres, locale)
    postgres_connections = _postgres_connection_rows(postgres, locale)
    rebuild_confirmation = _escape(json.dumps(translate(locale, "adminMaintenance.rebuildConfirm")))
    maintenance_notice = _admin_maintenance_notice(request, locale)
    raw_retention_rows = "".join(
        _admin_retention_row(locale, "raw-events", months, retention[months])
        for months in ADMIN_RETENTION_MONTHS
    )
    qso_retention_rows = "".join(
        _admin_retention_row(locale, "qsos", months, retention[months])
        for months in ADMIN_RETENTION_MONTHS
    )
    content = f"""
<section class="card"><div class="nav"><h1 style="margin:0">{_escape(translate(locale, "admin.title"))}</h1><form method="post" action="/admin/logout"><button class="button secondary" type="submit">{_escape(translate(locale, "admin.logout"))}</button></form></div>
<div class="stats"><div class="stat"><small>{_escape(translate(locale, "admin.registeredUsers"))}</small><strong>{_escape(stats['total_users'])}</strong></div><div class="stat"><small>{_escape(translate(locale, "admin.activeUsers"))}</small><strong>{_escape(stats['active_users'])}</strong></div><div class="stat"><small>{_escape(translate(locale, "admin.qso24"))}</small><strong>{_escape(stats['qso_count'])}</strong></div><div class="stat"><small>{_escape(translate(locale, "admin.talkTime24"))}</small><strong>{_escape(round(stats['duration_seconds'],1))} s</strong></div></div></section>
{maintenance_notice}
<section class="card"><h2>{_escape(translate(locale, "admin.registeredUsers"))}</h2><p class="muted">{_escape(translate(locale, "admin.userMetrics"))}</p><div class="table-wrap"><table><thead><tr><th>{_escape(translate(locale, "user.callsign"))}</th><th>{_escape(translate(locale, "user.name"))}</th><th>{_escape(translate(locale, "user.email"))}</th><th>{_escape(translate(locale, "admin.status"))}</th><th>{_escape(translate(locale, "user.qsoCount"))}</th><th>{_escape(translate(locale, "home.talkTime"))}</th><th>{_escape(translate(locale, "admin.actions"))}</th></tr></thead><tbody>{rows}</tbody></table></div></section>
<section class="card"><div class="nav"><h2 style="margin:0">{_escape(translate(locale, "admin.postgresql"))}</h2><form method="post" action="/admin/postgres/analyze"><button class="button secondary" type="submit">{_escape(translate(locale, "admin.refreshPlanner"))}</button></form></div>
<div class="stats"><div class="stat"><small>{_escape(translate(locale, "admin.database"))}</small><strong>{_escape(postgres['database_name'])}</strong></div><div class="stat"><small>{_escape(translate(locale, "admin.databaseSize"))}</small><strong>{_escape(postgres['database_size'])}</strong></div><div class="stat"><small>{_escape(translate(locale, "admin.activeConnections"))}</small><strong>{_escape(postgres['active_connections'])}</strong></div><div class="stat"><small>{_escape(translate(locale, "admin.totalConnections"))}</small><strong>{_escape(postgres['total_connections'])}</strong></div></div>
<p class="muted"><strong>{_escape(translate(locale, "admin.serverStarted"))}:</strong> {_escape(_format_datetime(postgres['server_started_at']))} · <strong>{_escape(translate(locale, "admin.serverTime"))}:</strong> {_escape(_format_datetime(postgres['server_time']))}</p>
<p class="muted"><strong>{_escape(translate(locale, "admin.version"))}:</strong> {_escape(postgres['version'])}</p>
<div class="charts"><div><h3>{_escape(translate(locale, "admin.applicationTables"))}</h3><div class="table-wrap"><table><thead><tr><th>{_escape(translate(locale, "admin.database"))}</th><th>{_escape(translate(locale, "admin.estimatedRows"))}</th><th>{_escape(translate(locale, "admin.totalSize"))}</th></tr></thead><tbody>{postgres_tables}</tbody></table></div></div><div><h3>{_escape(translate(locale, "admin.connectionStates"))}</h3><div class="table-wrap"><table><thead><tr><th>{_escape(translate(locale, "admin.state"))}</th><th>{_escape(translate(locale, "admin.connections"))}</th></tr></thead><tbody>{postgres_connections}</tbody></table></div></div></div></section>
<section class="card"><h2>{_escape(translate(locale, "adminMaintenance.title"))}</h2><p class="muted">{_escape(translate(locale, "adminMaintenance.description"))}</p>
<div class="card" style="background:#f8f9ff;box-shadow:none;margin:0 0 16px;padding:18px"><h3>{_escape(translate(locale, "adminMaintenance.rebuildTitle"))}</h3><p class="muted">{_escape(translate(locale, "adminMaintenance.rebuildDescription"))}</p><p><strong>{_escape(translate(locale, "adminMaintenance.currentQsos"))}:</strong> {_escape(maintenance['qsos'])} · <strong>{_escape(translate(locale, "adminMaintenance.rawEventsScanned"))}:</strong> {_escape(maintenance['raw_events'])}</p><form method="post" action="/admin/maintenance/rebuild-qsos" onsubmit="return window.confirm({rebuild_confirmation})"><button class="button" type="submit">{_escape(translate(locale, "adminMaintenance.rebuildButton"))}</button></form></div>
<h3>{_escape(translate(locale, "adminMaintenance.rawTitle"))}</h3>{raw_retention_rows}
<h3>{_escape(translate(locale, "adminMaintenance.qsoTitle"))}</h3>{qso_retention_rows}
</section>
"""
    return _account_page_with_metrics(
        translate(locale, "admin.title"),
        content,
        locale,
        records_retrieved=len(users),
        query_seconds=query_seconds,
    )


@app.get("/admin/stats")
def admin_stats(request: Request) -> JSONResponse:
    if not _admin_allowed(request):
        return JSONResponse({"error": "admin authentication required"}, status_code=401)
    return JSONResponse(jsonable_encoder(get_store().admin_statistics()))


@app.get("/admin/users")
def admin_users(request: Request) -> JSONResponse:
    if not _admin_allowed(request):
        return JSONResponse({"error": "admin authentication required"}, status_code=401)
    return JSONResponse(jsonable_encoder(get_store().list_users()))


@app.get("/admin/postgres")
def admin_postgres(request: Request) -> JSONResponse:
    if not _admin_allowed(request):
        return JSONResponse({"error": "admin authentication required"}, status_code=401)
    return JSONResponse(jsonable_encoder(get_store().postgres_overview()))


@app.post("/admin/postgres/analyze")
def admin_postgres_analyze(request: Request) -> Response:
    if not _admin_allowed(request):
        return JSONResponse({"error": "admin authentication required"}, status_code=401)
    get_store().analyze_postgres()
    if request.headers.get("accept", "").startswith("application/json"):
        return JSONResponse({"analyzed": True})
    return _admin_redirect()


@app.post("/admin/maintenance/rebuild-qsos")
def admin_rebuild_qsos(request: Request) -> Response:
    if not _admin_allowed(request):
        return JSONResponse({"error": "admin authentication required"}, status_code=401)
    result = get_store().rebuild_qsos_from_raw_events(settings.kerchunk_threshold_seconds)
    if request.headers.get("accept", "").startswith("application/json"):
        return JSONResponse(jsonable_encoder(result))
    return _admin_redirect(urlencode({
        "notice": "rebuild",
        "count": result["qsos_rebuilt"],
        "raw": result["raw_events_scanned"],
    }))


@app.post("/admin/maintenance/raw-events/{months}")
def admin_clear_raw_events(months: int, request: Request) -> Response:
    if not _admin_allowed(request):
        return JSONResponse({"error": "admin authentication required"}, status_code=401)
    if months not in ADMIN_RETENTION_MONTHS:
        return JSONResponse({"error": "retention period must be 1, 2, 3, or 6 months"}, status_code=400)
    result = get_store().clear_old_raw_events(months)
    if request.headers.get("accept", "").startswith("application/json"):
        return JSONResponse(jsonable_encoder(result))
    return _admin_redirect(urlencode({
        "notice": "raw",
        "months": months,
        "count": result["raw_events_deleted"],
        "qsos": result["qsos_deleted"],
    }))


@app.post("/admin/maintenance/qsos/{months}")
def admin_clear_qsos(months: int, request: Request) -> Response:
    if not _admin_allowed(request):
        return JSONResponse({"error": "admin authentication required"}, status_code=401)
    if months not in ADMIN_RETENTION_MONTHS:
        return JSONResponse({"error": "retention period must be 1, 2, 3, or 6 months"}, status_code=400)
    result = get_store().clear_old_qsos(months)
    if request.headers.get("accept", "").startswith("application/json"):
        return JSONResponse(jsonable_encoder(result))
    return _admin_redirect(urlencode({
        "notice": "qso",
        "months": months,
        "count": result["qsos_deleted"],
    }))


@app.post("/admin/users/{user_id}/status")
async def admin_user_status(user_id: int, request: Request) -> Response:
    if not _admin_allowed(request):
        return JSONResponse({"error": "admin authentication required"}, status_code=401)
    fields = await _form_fields(request)
    active = fields.get("active", "false").lower() in {"1", "true", "yes", "on"}
    updated = get_store().set_user_active(user_id, active)
    if request.headers.get("content-type", "").startswith("application/json"):
        return JSONResponse({"updated": updated, "active": active})
    return _admin_redirect()


@app.post("/admin/users/{user_id}/delete")
def admin_user_delete_page(user_id: int, request: Request) -> Response:
    if not _admin_allowed(request):
        return JSONResponse({"error": "admin authentication required"}, status_code=401)
    deleted = get_store().delete_user(user_id)
    if request.headers.get("accept", "").startswith("application/json"):
        return JSONResponse({"deleted": deleted})
    return _admin_redirect()


@app.delete("/admin/users/{user_id}")
def admin_user_delete(user_id: int, request: Request) -> JSONResponse:
    if not _admin_allowed(request):
        return JSONResponse({"error": "admin authentication required"}, status_code=401)
    return JSONResponse({"deleted": get_store().delete_user(user_id)})
