"""Authenticated reporting routes, charts, and export serializers."""

from __future__ import annotations

import csv
from datetime import datetime
import io
from time import perf_counter
from typing import Any
from urllib.parse import urlencode

from fastapi import APIRouter, Query, Request
from fastapi.responses import HTMLResponse, JSONResponse, RedirectResponse
from starlette.responses import Response

from . import web
from .i18n import catalog, translate


router = APIRouter()
TIME_RANGES = web.TIME_RANGES


def get_store():
    return web.get_store()


def _current_user(request):
    return web._current_user(request)


def request_locale(request):
    return web.request_locale(request)


def start_time(value):
    return web.start_time(value)


def end_time(value):
    return web.end_time(value)


def histogram_bucket_seconds(value):
    return web.histogram_bucket_seconds(value)


def _escape(value):
    return web._escape(value)


def _format_datetime(value):
    return web._format_datetime(value)


def _account_page_with_metrics(*args, **kwargs):
    return web._account_page_with_metrics(*args, **kwargs)


def _report_time_range(value: str | None) -> str:
    return value if value in TIME_RANGES else "1M"


def _report_query(
    time_range: str,
    continent: str | None,
    country: str | None,
    talkgroups: list[int] | None,
    callsign: str | None,
) -> str:
    values: list[tuple[str, str]] = [("timeRange", time_range)]
    if callsign:
        values.append(("callsign", callsign))
    if continent:
        values.append(("continent", continent))
    if country:
        values.append(("country", country))
    for talkgroup in talkgroups or []:
        values.append(("talkgroup", str(talkgroup)))
    return urlencode(values)


def _report_duration(seconds: float | int) -> str:
    seconds = round(float(seconds or 0))
    hours, remainder = divmod(seconds, 3600)
    minutes, seconds = divmod(remainder, 60)
    if hours:
        return f"{hours}:{minutes:02d}:{seconds:02d}"
    if minutes:
        return f"{minutes}:{seconds:02d}"
    return f"{seconds} s"


def _admin_seconds(value: Any) -> str:
    if value is None:
        return "—"
    return f"{float(value):.1f} s"


def _admin_hms(value: Any) -> str:
    total_seconds = max(0, round(float(value or 0)))
    hours, remainder = divmod(total_seconds, 3600)
    minutes, seconds = divmod(remainder, 60)
    return f"{hours}:{minutes:02d}:{seconds:02d}"


def _report_histogram_label(value: Any, bucket_seconds: int) -> str:
    """Return a compact label that fits beneath a report histogram bar."""
    parsed = value
    if isinstance(value, str):
        try:
            parsed = datetime.fromisoformat(value.replace("Z", "+00:00"))
        except ValueError:
            parsed = None
    if isinstance(parsed, datetime):
        return parsed.strftime("%d/%m" if bucket_seconds >= 24 * 60 * 60 else "%H:%M")
    return str(value or "")[:8]


def _report_callsign_chart_rows(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    result = []
    for row in rows:
        callsign = str(row.get("callsign") or "—")
        name = str(row.get("source_name") or "").strip()
        result.append({**row, "report_label": f"{callsign} · {name}" if name else callsign})
    return result


REPORT_ENTRY_LIMIT = 50


def _limit_report_entries(report: dict[str, Any]) -> dict[str, Any]:
    """Keep report tables and exports bounded while retaining the full histogram."""
    limited = dict(report)
    limited["daily"] = list(report.get("daily", []))[:REPORT_ENTRY_LIMIT]
    limited["talkgroups"] = sorted(
        report.get("talkgroups", []),
        key=lambda row: (-int(row.get("qso_count") or 0), str(row.get("name") or "")),
    )[:REPORT_ENTRY_LIMIT]
    limited["callsigns"] = sorted(
        report.get("callsigns", []),
        key=lambda row: (-int(row.get("qso_count") or 0), str(row.get("callsign") or "")),
    )[:REPORT_ENTRY_LIMIT]
    return limited


def _report_bar_rows(rows: list[dict[str, Any]], label: str, value: str, formatter: Any, empty_label: str = "No data for this filter.") -> str:
    if not rows:
        return f'<p class="muted">{_escape(empty_label)}</p>'
    maximum = max(float(row.get(value) or 0) for row in rows) or 1
    result = []
    for row in rows:
        amount = float(row.get(value) or 0)
        width = max(2, amount / maximum * 100)
        result.append(
            f'<div class="report-bar"><span>{_escape(row.get(label, "—"))}</span>'
            f'<div><i style="width:{width:.1f}%"></i></div><strong>{_escape(formatter(amount))}</strong></div>'
        )
    return "".join(result)


def _report_histogram(rows: list[dict[str, Any]], bucket_seconds: int, locale: str, empty_label: str) -> str:
    if not rows or not any(int(row.get("qso_count") or 0) for row in rows):
        return f'<p class="muted">{_escape(empty_label)}</p>'
    maximum = max(int(row.get("qso_count") or 0) for row in rows) or 1
    columns = []
    label_step = max(1, len(rows) // 10)
    for index, row in enumerate(rows):
        bucket = row.get("bucket")
        label = _report_histogram_label(bucket, bucket_seconds)
        count = int(row.get("qso_count") or 0)
        height = max(2, count / maximum * 100) if count else 0
        visible_label = label if index % label_step == 0 or index == len(rows) - 1 else ""
        columns.append(
            f'<div class="histogram-column" title="{_escape(label)}: {_escape(count)} '
            f'{_escape(translate(locale, "reports.qsos"))}">'
            f'<span class="histogram-value">{_escape(count) if count else ""}</span>'
            f'<i style="height:{height:.1f}%"></i>'
            f'<small>{_escape(visible_label)}</small></div>'
        )
    return f'<div class="histogram">{"".join(columns)}</div>'


def _report_hour_rows(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    return [
        {**row, "label": f"{int(row.get('hour') or 0):02d}:00"}
        for row in rows
    ]


def _report_weekday_rows(rows: list[dict[str, Any]], locale: str) -> list[dict[str, Any]]:
    weekday_keys = {
        1: "monday",
        2: "tuesday",
        3: "wednesday",
        4: "thursday",
        5: "friday",
        6: "saturday",
        7: "sunday",
    }
    return [
        {
            **row,
            "label": translate(locale, f"reports.{weekday_keys.get(int(row.get('weekday') or 0), 'monday')}"),
        }
        for row in rows
    ]


def _report_concurrency_chart(
    rows: list[dict[str, Any]],
    bucket_seconds: int,
    locale: str,
    empty_label: str,
) -> str:
    if not rows or not any(
        int(row.get("active_talkgroups") or 0) or int(row.get("active_sources") or 0)
        for row in rows
    ):
        return f'<p class="muted">{_escape(empty_label)}</p>'
    maximum = max(
        max(int(row.get("active_talkgroups") or 0), int(row.get("active_sources") or 0))
        for row in rows
    ) or 1
    label_step = max(1, len(rows) // 10)
    columns = []
    for index, row in enumerate(rows):
        label = _report_histogram_label(row.get("bucket"), bucket_seconds)
        visible_label = label if index % label_step == 0 or index == len(rows) - 1 else ""
        talkgroups = int(row.get("active_talkgroups") or 0)
        sources = int(row.get("active_sources") or 0)
        columns.append(
            f'<div class="concurrency-column" title="{_escape(label)}">'
            f'<i class="concurrency-talkgroups" style="height:{talkgroups / maximum * 100:.1f}%" '
            f'title="{_escape(translate(locale, "reports.activeTalkgroups"))}: {talkgroups}"></i>'
            f'<i class="concurrency-sources" style="height:{sources / maximum * 100:.1f}%" '
            f'title="{_escape(translate(locale, "reports.activeSources"))}: {sources}"></i>'
            f'<span class="concurrency-value" style="position:absolute;top:-16px;font-size:9px;color:#6b7280;white-space:nowrap">{talkgroups}/{sources}</span>'
            f'<small>{_escape(visible_label)}</small></div>'
        )
    legend = (
        f'<div class="concurrency-legend">'
        f'<span><i class="concurrency-talkgroups"></i>{_escape(translate(locale, "reports.activeTalkgroups"))}</span>'
        f'<span><i class="concurrency-sources"></i>{_escape(translate(locale, "reports.activeSources"))}</span>'
        f'</div>'
    )
    return f'{legend}<div class="concurrency-chart">{"".join(columns)}</div>'


def _report_page(
    request: Request,
    user: dict[str, Any],
    time_range: str,
    continent: str | None,
    country: str | None,
    talkgroups: list[int] | None,
    callsign: str | None,
) -> HTMLResponse:
    locale = request_locale(request)
    query_started = perf_counter()
    store = get_store()
    report_start = start_time(time_range)
    report_end = end_time(time_range)
    if report_end is None:
        report = store.user_report(
            callsign,
            report_start,
            continent,
            country,
            talkgroups,
            histogram_bucket_seconds(time_range),
        )
    else:
        report = store.user_report(
            callsign,
            report_start,
            continent,
            country,
            talkgroups,
            histogram_bucket_seconds(time_range),
            end_time=report_end,
        )
    if hasattr(store, "active_talkgroups"):
        talkgroup_selector_rows = [
            {
                "talkgroup_id": row["value"],
                "name": row["label"],
                "qso_count": row["count"],
            }
            for row in (
                store.active_talkgroups(
                    start_time(time_range), continent, country
                )
                if report_end is None
                else store.active_talkgroups(
                    report_start, continent, country, end_time=report_end
                )
            )
        ]
    else:
        talkgroup_selector_rows = list(report.get("talkgroups", []))
    report = _limit_report_entries(report)
    translations = catalog(locale)
    metadata = translations.get("metadata", {})
    country_labels = metadata.get("countries", {})
    continents = get_store().continents()
    countries = sorted(
        get_store().countries(continent),
        key=lambda row: str(country_labels.get(row["value"], row["label"])).casefold(),
    )
    query_seconds = perf_counter() - query_started
    selected_talkgroups = {str(value) for value in talkgroups or []}
    range_keys = {
        "5m": "home.last5m", "15m": "home.last15m", "30m": "home.last30m",
        "1h": "home.last1h", "2h": "home.last2h", "6h": "home.last6h",
        "12h": "home.last12h", "24h": "home.last24h", "today": "home.today", "yesterday": "home.yesterday",
        "2d": "home.last2d", "5d": "home.last5d", "1w": "home.last1w",
        "lastWeek": "home.lastWeek", "2w": "home.last2w", "1M": "home.last1M",
        "lastMonth": "home.lastMonth", "2M": "home.last2M", "3M": "home.last3M",
    }
    range_options = "".join(
        f'<option value="{key}"{" selected" if key == time_range else ""}>'
        f'{_escape(translate(locale, label_key, key))}</option>'
        for key, label_key in range_keys.items()
    )
    continent_options = "".join(
        f'<option value="{_escape(value)}"{" selected" if value == continent else ""}>'
        f'{_escape(metadata.get("continents", {}).get(value, value))}</option>'
        for value in continents
    )
    country_options = "".join(
        f'<option value="{_escape(row["value"])}"{" selected" if row["value"] == country else ""}>'
        f'{_escape(country_labels.get(row["value"], row["label"]))}</option>'
        for row in countries
    )
    report["talkgroups"] = sorted(
        report["talkgroups"],
        key=lambda row: (-int(row.get("qso_count") or 0), str(row.get("name") or ""), int(row.get("talkgroup_id") or 0)),
    )
    talkgroup_selector_rows.sort(
        key=lambda row: (
            -int(row.get("qso_count") or 0),
            str(row.get("name") or "").casefold(),
            int(row.get("talkgroup_id") or 0),
        )
    )
    selected_specific = any(
        str(row.get("talkgroup_id")) in selected_talkgroups
        for row in talkgroup_selector_rows
    )
    all_talkgroups_selected = not selected_specific
    talkgroup_options = (
        f'<label class="report-talkgroup-option"><input type="checkbox" '
        f'data-all-talkgroups value="all"{" checked" if all_talkgroups_selected else ""}>'
        f'<span>{_escape(translate(locale, "home.allTalkgroups"))}</span></label>'
        + "".join(
            f'<label class="report-talkgroup-option"><input type="checkbox" '
            f'data-talkgroup-value="{_escape(row["talkgroup_id"])}" '
            f'{" checked" if not all_talkgroups_selected and str(row["talkgroup_id"]) in selected_talkgroups else ""}>'
            f'<span>{_escape(row.get("name"))} ({_escape(row.get("qso_count"))} '
            f'{_escape(translate(locale, "reports.qsos"))})</span></label>'
            for row in talkgroup_selector_rows
        )
        + '<select id="reportTalkgroupValues" name="talkgroup" multiple hidden>'
        + "".join(
            f'<option value="{_escape(row["talkgroup_id"])}"'
            f'{" selected" if not all_talkgroups_selected and str(row["talkgroup_id"]) in selected_talkgroups else ""}>'
            f'{_escape(row.get("name"))}</option>'
            for row in talkgroup_selector_rows
        )
        + '</select>'
    )
    query = _report_query(time_range, continent, country, talkgroups, callsign)
    report_scope = callsign.strip() if callsign and callsign.strip() else translate(locale, "reports.allCallsigns")
    no_data = translate(locale, "reports.noData")
    summary = report["summary"]
    hourly_rows = _report_hour_rows(report.get("hourly_activity", []))
    weekday_rows = _report_weekday_rows(report.get("weekday_activity", []), locale)
    peak_rows = [
        {
            **row,
            "label": _report_histogram_label(row.get("bucket"), histogram_bucket_seconds(time_range)),
        }
        for row in report.get("peak_periods", [])
    ]
    traffic_trend = report.get("traffic_trend", {})
    trend_change = traffic_trend.get("qso_change_percent")
    trend_change_label = (
        translate(locale, "reports.noPreviousData")
        if trend_change is None
        else f"{trend_change:+.1f}%"
    )
    growth_rows = report.get("talkgroup_growth", [])
    concurrency_rows = report.get("concurrent_activity", [])
    daily_rows = "".join(
        f'<tr><td>{_escape(row["day"])}</td><td>{_escape(row["qso_count"])}</td>'
        f'<td>{_escape(_report_duration(row["duration_seconds"]))}</td></tr>'
        for row in report["daily"]
    ) or f'<tr><td colspan="3" class="muted">{_escape(no_data)}</td></tr>'
    talkgroup_rows = "".join(
        f'<tr><td>{_escape(row["name"])}</td><td>{_escape(row["talkgroup_id"])}</td>'
        f'<td>{_escape(row["qso_count"])}</td><td>{_escape(_report_duration(row["duration_seconds"]))}</td>'
        f'<td>{_escape(_format_datetime(row["last_seen_at"]))}</td></tr>'
        for row in report["talkgroups"]
    ) or f'<tr><td colspan="5" class="muted">{_escape(no_data)}</td></tr>'
    callsign_chart_rows = _report_callsign_chart_rows(report["callsigns"])
    callsign_rows = "".join(
        f'<tr><td>{_escape(row["callsign"])}</td>'
        f'<td>{_escape(row.get("source_name") or "—")}</td>'
        f'<td>{_escape(row.get("countries") or "—")}</td>'
        f'<td>{_escape(row["qso_count"])}</td>'
        f'<td>{_escape(_report_duration(row["duration_seconds"]))}</td>'
        f'<td>{_escape(row["unique_talkgroups"])}</td></tr>'
        for row in report["callsigns"]
    ) or f'<tr><td colspan="6" class="muted">{_escape(no_data)}</td></tr>'
    content = f"""
<style>.report-controls{{display:flex;gap:12px;flex-wrap:wrap;align-items:flex-start}}.report-controls label{{display:block;font-size:12px;font-weight:700;color:#6b7280;margin-bottom:5px}}.report-controls select,.report-controls input,.report-controls button{{height:40px;padding:0 10px;border:2px solid #e5e7eb;border-radius:8px;background:#fff;font:inherit}}.report-controls input{{min-width:180px}}.report-controls button{{margin-top:21px;border:0;background:linear-gradient(135deg,#667eea,#764ba2);color:#fff;font-weight:700;cursor:pointer}}.report-talkgroup-control{{min-width:250px}}.report-talkgroup-list{{display:flex;flex-direction:column;gap:4px;max-height:150px;min-width:250px;overflow-y:auto;padding:7px 9px;border:2px solid #e5e7eb;border-radius:8px;background:#fff}}.report-talkgroup-option{{display:flex!important;align-items:center;gap:7px;margin:0!important;color:#1f2937;font-size:13px!important;font-weight:500!important;white-space:nowrap}}.report-talkgroup-option input{{width:auto;height:auto;margin:0;accent-color:#667eea}}.report-actions{{display:flex;gap:9px;flex-wrap:wrap;margin-top:16px}}.report-bar{{display:grid;grid-template-columns:220px 1fr 75px;gap:8px;align-items:center;margin:8px 0;font-size:12px}}.report-bar span{{overflow:hidden;text-overflow:ellipsis;white-space:nowrap}}.report-bar div{{height:18px;background:#f0f1f6;border-radius:5px;overflow:hidden}}.report-bar i{{display:block;height:100%;background:linear-gradient(90deg,#667eea,#764ba2);border-radius:5px}}.report-bar strong{{text-align:right;color:#6b7280;white-space:nowrap}}.report-callsign-table th,.report-callsign-table td{{white-space:nowrap}}.histogram{{height:240px;display:flex;align-items:end;gap:3px;padding:18px 4px 28px;border-bottom:1px solid #e5e7eb;overflow:hidden}}.histogram-column{{position:relative;display:flex;flex:1;min-width:8px;height:100%;align-items:end;justify-content:end}}.histogram-column i{{display:block;width:100%;min-height:0;background:linear-gradient(180deg,#764ba2,#667eea);border-radius:4px 4px 0 0}}.histogram-column small{{position:absolute;bottom:-24px;left:50%;transform:translateX(-50%);width:56px;max-width:56px;overflow:hidden;text-overflow:ellipsis;font-size:9px;color:#6b7280;text-align:center;white-space:nowrap}}.histogram-value{{position:absolute;top:-16px;font-size:10px;color:#6b7280}}.concurrency-legend{{display:flex;gap:18px;margin:4px 0 8px;font-size:12px;color:#6b7280}}.concurrency-legend span{{display:flex;align-items:center;gap:5px}}.concurrency-legend i{{display:inline-block;width:10px;height:10px;border-radius:2px}}.concurrency-chart{{height:240px;display:flex;align-items:end;gap:3px;padding:18px 4px 28px;border-bottom:1px solid #e5e7eb;overflow:hidden}}.concurrency-column{{position:relative;display:flex;flex:1;min-width:8px;height:100%;align-items:end;justify-content:center;gap:1px}}.concurrency-column i{{display:block;width:calc(50% - 1px);min-height:0;border-radius:4px 4px 0 0}}.concurrency-column small{{position:absolute;bottom:-24px;left:50%;transform:translateX(-50%);width:56px;max-width:56px;overflow:hidden;text-overflow:ellipsis;font-size:9px;color:#6b7280;text-align:center;white-space:nowrap}}.concurrency-talkgroups{{background:#667eea}}.concurrency-sources{{background:#764ba2}}@media(max-width:700px){{.report-controls{{flex-direction:column}}.report-talkgroup-list{{min-width:0;width:100%}}.report-bar{{grid-template-columns:160px 1fr 55px}}.histogram,.concurrency-chart{{gap:2px}}.histogram-value{{display:none}}}}</style>
<section class="card"><div class="nav"><div><h1 style="margin:0;text-align:left">{_escape(translate(locale, "reports.title"))}</h1><p class="muted">{_escape(translate(locale, "reports.forUser"))}: {_escape(user['callsign'])} · {_escape(translate(locale, "reports.scope"))}: {_escape(report_scope)}</p></div><div><a class="button secondary" href="/user/live-qsos">{_escape(translate(locale, "live.title"))}</a> <form class="inline" method="post" action="/user/logout"><button class="button secondary" type="submit">{_escape(translate(locale, "user.logout"))}</button></form></div></div>
<form id="reportForm" class="report-controls" method="get" action="/user/reports"><div><label>{_escape(translate(locale, "reports.callsign"))}</label><input name="callsign" value="{_escape(callsign or '')}" placeholder="{_escape(translate(locale, "home.callsignPlaceholder"))}"></div><div><label>{_escape(translate(locale, "reports.timeRange"))}</label><select id="reportTimeRange" name="timeRange">{range_options}</select></div><div><label>{_escape(translate(locale, "home.continent"))}</label><select id="reportContinent" name="continent"><option value="">{_escape(translate(locale, "home.allContinents"))}</option>{continent_options}</select></div><div><label>{_escape(translate(locale, "home.country"))}</label><select id="reportCountry" name="country"><option value="">{_escape(translate(locale, "home.allCountries"))}</option>{country_options}</select></div><div class="report-talkgroup-control"><label>{_escape(translate(locale, "reports.talkgroups"))}</label><div id="reportTalkgroups" class="report-talkgroup-list" role="group" name="talkgroup" multiple>{talkgroup_options}</div></div><button type="submit">{_escape(translate(locale, "reports.generate"))}</button></form>
<div class="report-actions"><a class="button secondary" href="/user/reports/export.csv?{query}">{_escape(translate(locale, "reports.csv"))}</a><a class="button secondary" href="/user/reports/export.xlsx?{query}">{_escape(translate(locale, "reports.excel"))}</a><a class="button" href="/user/reports/export.pdf?{query}">{_escape(translate(locale, "reports.pdf"))}</a></div></section>
<section class="card"><h2>{_escape(translate(locale, "reports.summary"))}</h2><div class="stats"><div class="stat"><small>{_escape(translate(locale, "reports.qsos"))}</small><strong>{_escape(summary['qso_count'])}</strong></div><div class="stat"><small>{_escape(translate(locale, "reports.talkTime"))}</small><strong>{_escape(_report_duration(summary['duration_seconds']))}</strong></div><div class="stat"><small>{_escape(translate(locale, "reports.uniqueTalkgroups"))}</small><strong>{_escape(summary['unique_talkgroups'])}</strong></div><div class="stat"><small>{_escape(translate(locale, "reports.activeDays"))}</small><strong>{_escape(summary['active_days'])}</strong></div></div></section>
<section class="card"><h2>{_escape(translate(locale, "reports.dailyActivity"))}</h2>{_report_histogram(report['histogram'], histogram_bucket_seconds(time_range), locale, no_data)}</section>
<section class="card"><h2>{_escape(translate(locale, "reports.trafficTrend"))}</h2><div class="stats"><div class="stat"><small>{_escape(translate(locale, "reports.currentPeriod"))}</small><strong>{_escape(traffic_trend.get("current_qso_count", summary["qso_count"]))} { _escape(translate(locale, "reports.qsos")) }</strong></div><div class="stat"><small>{_escape(translate(locale, "reports.previousPeriod"))}</small><strong>{_escape(traffic_trend.get("previous_qso_count", 0))} { _escape(translate(locale, "reports.qsos")) }</strong></div><div class="stat"><small>{_escape(translate(locale, "reports.change"))}</small><strong>{_escape(trend_change_label)}</strong></div></div></section>
<section class="card"><h2>{_escape(translate(locale, "reports.hourlyActivity"))}</h2>{_report_bar_rows(hourly_rows, 'label', 'qso_count', lambda value: f'{int(value)} {translate(locale, "reports.qsos")}', no_data)}</section>
<section class="card"><h2>{_escape(translate(locale, "reports.weekdayActivity"))}</h2>{_report_bar_rows(weekday_rows, 'label', 'qso_count', lambda value: f'{int(value)} {translate(locale, "reports.qsos")}', no_data)}</section>
<section class="card"><h2>{_escape(translate(locale, "reports.peakPeriods"))}</h2>{_report_bar_rows(peak_rows, 'label', 'qso_count', lambda value: f'{int(value)} {translate(locale, "reports.qsos")}', no_data)}</section>
<section class="card"><h2>{_escape(translate(locale, "reports.concurrentActivity"))}</h2>{_report_concurrency_chart(concurrency_rows, histogram_bucket_seconds(time_range), locale, no_data)}</section>
<section class="card"><h2>{_escape(translate(locale, "reports.talkgroupActivity"))}</h2>{_report_bar_rows(report['talkgroups'], 'name', 'qso_count', lambda value: f'{int(value)} {translate(locale, "reports.qsos")}', no_data)}<div class="table-wrap"><table><thead><tr><th>{_escape(translate(locale, "home.talkgroup"))}</th><th>{_escape(translate(locale, "user.id"))}</th><th>{_escape(translate(locale, "reports.qsos"))}</th><th>{_escape(translate(locale, "reports.talkTime"))}</th><th>{_escape(translate(locale, "home.lastHeard"))}</th></tr></thead><tbody>{talkgroup_rows}</tbody></table></div></section>
<section class="card"><h2>{_escape(translate(locale, "reports.talkgroupGrowth"))}</h2><div class="table-wrap"><table><thead><tr><th>{_escape(translate(locale, "home.talkgroup"))}</th><th>{_escape(translate(locale, "reports.currentPeriod"))}</th><th>{_escape(translate(locale, "reports.previousPeriod"))}</th><th>{_escape(translate(locale, "reports.growth"))}</th></tr></thead><tbody>{''.join(f'<tr><td>{_escape(row.get("name"))}</td><td>{_escape(row.get("current_qso_count"))}</td><td>{_escape(row.get("previous_qso_count"))}</td><td>{_escape(translate(locale, "reports.new")) if row.get("growth_percent") is None else _escape(f"{row["growth_percent"]:+.1f}%")}</td></tr>' for row in growth_rows) or f'<tr><td colspan="4" class="muted">{_escape(no_data)}</td></tr>'}</tbody></table></div></section>
<section class="card"><h2>{_escape(translate(locale, "reports.callsignActivity"))}</h2>{_report_bar_rows(callsign_chart_rows, 'report_label', 'qso_count', lambda value: f'{int(value)} {translate(locale, "reports.qsos")}', no_data)}<div class="table-wrap"><table class="report-callsign-table"><thead><tr><th>{_escape(translate(locale, "home.callsignFilter"))}</th><th>{_escape(translate(locale, "user.name"))}</th><th>{_escape(translate(locale, "home.country"))}</th><th>{_escape(translate(locale, "reports.qsos"))}</th><th>{_escape(translate(locale, "reports.talkTime"))}</th><th>{_escape(translate(locale, "reports.uniqueTalkgroups"))}</th></tr></thead><tbody>{callsign_rows}</tbody></table></div></section>
<section class="card"><h2>{_escape(translate(locale, "reports.dailyTable"))}</h2><div class="table-wrap"><table><thead><tr><th>{_escape(translate(locale, "reports.date"))}</th><th>{_escape(translate(locale, "reports.qsos"))}</th><th>{_escape(translate(locale, "reports.talkTime"))}</th></tr></thead><tbody>{daily_rows}</tbody></table></div></section>
<script>const reportCountry=document.getElementById('reportCountry');if(reportCountry){{const options=Array.from(reportCountry.options).slice(1).sort((a,b)=>new Intl.Collator(document.documentElement.lang||'en',{{sensitivity:'base',numeric:true}}).compare(a.textContent,b.textContent));options.forEach(option=>reportCountry.appendChild(option));}}const reportTalkgroups=document.getElementById('reportTalkgroups'),reportTalkgroupValues=document.getElementById('reportTalkgroupValues');const syncReportTalkgroups=()=>{{if(!reportTalkgroupValues)return;const selected=new Set(Array.from(reportTalkgroups.querySelectorAll('input[data-talkgroup-value]:checked')).map(input=>input.dataset.talkgroupValue));Array.from(reportTalkgroupValues.options).forEach(option=>option.selected=selected.has(option.value));}};reportTalkgroups?.addEventListener('change',event=>{{const input=event.target;if(input?.type!=='checkbox')return;const all=reportTalkgroups.querySelector('input[data-all-talkgroups]');const specifics=Array.from(reportTalkgroups.querySelectorAll('input[data-talkgroup-value]'));if(input.dataset.allTalkgroups!==undefined&&input.checked)specifics.forEach(item=>item.checked=false);else if(input.dataset.talkgroupValue!==undefined&&input.checked)all.checked=false;if(!Array.from(reportTalkgroups.querySelectorAll('input[type="checkbox"]')).some(item=>item.checked))all.checked=true;syncReportTalkgroups();}});syncReportTalkgroups();document.getElementById('reportContinent')?.addEventListener('change',()=>document.getElementById('reportForm').submit());document.getElementById('reportCountry')?.addEventListener('change',()=>document.getElementById('reportForm').submit());document.getElementById('reportTimeRange')?.addEventListener('change',()=>document.getElementById('reportForm').submit());</script>
"""
    return _account_page_with_metrics(
        translate(locale, "reports.title"),
        content,
        locale,
        records_retrieved=report["summary"]["qso_count"],
        query_seconds=query_seconds,
        user=user,
    )


def _load_user_report(
    request: Request,
    time_range: str,
    continent: str | None,
    country: str | None,
    talkgroups: list[int] | None,
    callsign: str | None,
) -> tuple[dict[str, Any] | None, dict[str, Any] | None]:
    user = _current_user(request)
    if user is None:
        return None, None
    report_start = start_time(time_range)
    report_end = end_time(time_range)
    if report_end is None:
        report = get_store().user_report(
            callsign,
            report_start,
            continent,
            country,
            talkgroups,
            histogram_bucket_seconds(time_range),
        )
    else:
        report = get_store().user_report(
            callsign,
            report_start,
            continent,
            country,
            talkgroups,
            histogram_bucket_seconds(time_range),
            end_time=report_end,
        )
    return user, _limit_report_entries(report)


def _report_csv(report: dict[str, Any]) -> bytes:
    report = _limit_report_entries(report)
    output = io.StringIO(newline="")
    writer = csv.writer(output)
    summary = report["summary"]
    writer.writerow(["Metric", "Value"])
    writer.writerow(["QSOs", summary["qso_count"]])
    writer.writerow(["Talk time seconds", summary["duration_seconds"]])
    writer.writerow(["Unique talkgroups", summary["unique_talkgroups"]])
    writer.writerow(["Active days", summary["active_days"]])
    writer.writerow([])
    writer.writerow(["Callsign", "Name", "Country", "QSOs", "Talk time seconds", "Unique talkgroups"])
    for row in report["callsigns"]:
        writer.writerow([
            row["callsign"], row.get("source_name") or "", row.get("countries") or "",
            row["qso_count"], row["duration_seconds"], row["unique_talkgroups"],
        ])
    return output.getvalue().encode("utf-8-sig")


def _report_excel(report: dict[str, Any]) -> bytes:
    from openpyxl import Workbook

    report = _limit_report_entries(report)
    workbook = Workbook()
    summary_sheet = workbook.active
    summary_sheet.title = "Summary"
    summary = report["summary"]
    summary_sheet.append(["Metric", "Value"])
    for label, value in (
        ("QSOs", summary["qso_count"]),
        ("Talk time seconds", summary["duration_seconds"]),
        ("Unique talkgroups", summary["unique_talkgroups"]),
        ("Active days", summary["active_days"]),
        ("First QSO", summary["first_qso_at"]),
        ("Last QSO", summary["last_qso_at"]),
    ):
        if isinstance(value, datetime) and value.tzinfo:
            value = value.replace(tzinfo=None)
        summary_sheet.append([label, value])
    daily_sheet = workbook.create_sheet("Daily activity")
    daily_sheet.append(["Date", "QSOs", "Talk time seconds"])
    for row in report["daily"]:
        daily_sheet.append([row["day"], row["qso_count"], row["duration_seconds"]])
    talkgroup_sheet = workbook.create_sheet("Talkgroups")
    talkgroup_sheet.append(["Talkgroup", "ID", "QSOs", "Talk time seconds", "Last seen"])
    for row in report["talkgroups"]:
        last_seen = row["last_seen_at"]
        if isinstance(last_seen, datetime) and last_seen.tzinfo:
            last_seen = last_seen.replace(tzinfo=None)
        talkgroup_sheet.append([row["name"], row["talkgroup_id"], row["qso_count"], row["duration_seconds"], last_seen])
    callsign_sheet = workbook.create_sheet("Callsigns")
    callsign_sheet.append(["Callsign", "Name", "Country", "QSOs", "Talk time seconds", "Unique talkgroups", "Last seen"])
    for row in report["callsigns"]:
        last_seen = row["last_seen_at"]
        if isinstance(last_seen, datetime) and last_seen.tzinfo:
            last_seen = last_seen.replace(tzinfo=None)
        callsign_sheet.append([
            row["callsign"], row.get("source_name") or "", row.get("countries") or "",
            row["qso_count"], row["duration_seconds"], row["unique_talkgroups"], last_seen,
        ])
    for sheet in workbook.worksheets:
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions
        for cell in sheet[1]:
            cell.font = cell.font.copy(bold=True)
    output = io.BytesIO()
    workbook.save(output)
    return output.getvalue()


def _pdf_histogram(
    reportlab_rows: list[dict[str, Any]],
    width: float,
    bucket_seconds: int = 86400,
) -> Any | None:
    from reportlab.lib.colors import HexColor
    from reportlab.graphics.shapes import Drawing, Line, Rect, String

    rows = reportlab_rows
    if not rows or not any(int(row.get("qso_count") or 0) for row in rows):
        return None
    height = 128
    left, bottom, chart_width, chart_height = 30, 28, width - 42, 86
    maximum = max(int(row.get("qso_count") or 0) for row in rows) or 1
    drawing = Drawing(width, height)
    drawing.add(Line(left, bottom, left + chart_width, bottom, strokeColor=HexColor("#9ca3af")))
    bar_width = chart_width / max(len(rows), 1)
    label_step = max(1, len(rows) // 8)
    for index, row in enumerate(rows):
        count = int(row.get("qso_count") or 0)
        x = left + index * bar_width + 0.5
        bar_height = count / maximum * chart_height if count else 0
        drawing.add(Rect(x, bottom, max(1, bar_width - 1), bar_height, fillColor=HexColor("#667eea"), strokeColor=None))
        if index % label_step == 0 or index == len(rows) - 1:
            label = _report_histogram_label(row.get("bucket"), bucket_seconds)
            drawing.add(String(x, 10, label, fontSize=6, fillColor="#6b7280"))
    return drawing


def _pdf_bar_chart(rows: list[dict[str, Any]], label: str, value: str, width: float) -> Any | None:
    from reportlab.lib.colors import HexColor
    from reportlab.graphics.shapes import Drawing, Rect, String

    rows = rows[:10]
    if not rows:
        return None
    left, row_height, bar_width = 115, 20, width - 150
    height = row_height * len(rows) + 6
    maximum = max(float(row.get(value) or 0) for row in rows) or 1
    drawing = Drawing(width, height)
    for index, row in enumerate(rows):
        y = height - (index + 1) * row_height + 4
        text = str(row.get(label, "—"))
        if len(text) > 19:
            text = text[:18] + "…"
        amount = float(row.get(value) or 0)
        drawing.add(String(0, y + 2, text, fontSize=7, fillColor=HexColor("#374151")))
        drawing.add(Rect(left, y, max(1, amount / maximum * bar_width), 10, fillColor=HexColor("#764ba2"), strokeColor=None))
        drawing.add(String(width - 30, y + 2, str(int(amount)), fontSize=7, fillColor=HexColor("#6b7280")))
    return drawing


def _report_pdf(
    report: dict[str, Any],
    callsign: str,
    locale: str,
    bucket_seconds: int = 86400,
) -> bytes:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.lib.units import mm
    from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

    report = _limit_report_entries(report)
    output = io.BytesIO()
    document = SimpleDocTemplate(output, pagesize=A4, rightMargin=14 * mm, leftMargin=14 * mm, topMargin=14 * mm, bottomMargin=14 * mm)
    styles = getSampleStyleSheet()
    story: list[Any] = [
        Paragraph(_escape(translate(locale, "reports.title")), styles["Title"]),
        Paragraph(f"{_escape(translate(locale, 'reports.forUser'))}: {_escape(callsign)}", styles["Normal"]),
        Spacer(1, 8),
    ]
    summary = report["summary"]
    story.append(Table(
        [[translate(locale, "reports.qsos"), summary["qso_count"], translate(locale, "reports.talkTime"), _report_duration(summary["duration_seconds"])],
         [translate(locale, "reports.uniqueTalkgroups"), summary["unique_talkgroups"], translate(locale, "reports.activeDays"), summary["active_days"]]],
        colWidths=[38 * mm, 32 * mm, 42 * mm, 45 * mm],
        style=TableStyle([("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#f1f3ff")), ("GRID", (0, 0), (-1, -1), 0.25, colors.lightgrey), ("FONTNAME", (0, 0), (-1, -1), "Helvetica"), ("VALIGN", (0, 0), (-1, -1), "MIDDLE")]),
    ))
    story.append(Spacer(1, 10))
    story.append(Paragraph(_escape(translate(locale, "reports.dailyActivity")), styles["Heading2"]))
    histogram = _pdf_histogram(report["histogram"], 170 * mm, bucket_seconds)
    if histogram is not None:
        story.append(histogram)
    story.append(Spacer(1, 8))
    story.append(Paragraph(_escape(translate(locale, "reports.talkgroupActivity")), styles["Heading2"]))
    talkgroup_chart = _pdf_bar_chart(report["talkgroups"], "name", "qso_count", 170 * mm)
    if talkgroup_chart is not None:
        story.append(talkgroup_chart)
        story.append(Spacer(1, 6))
    talkgroup_table = [[translate(locale, "home.talkgroup"), translate(locale, "user.id"), translate(locale, "reports.qsos"), translate(locale, "reports.talkTime")]] + [
        [str(row["name"]), str(row["talkgroup_id"]), str(row["qso_count"]), _report_duration(row["duration_seconds"])]
        for row in report["talkgroups"]
    ]
    if len(talkgroup_table) == 1:
        talkgroup_table.append([translate(locale, "reports.noData"), "", "", ""])
    story.append(Table(talkgroup_table, repeatRows=1, style=TableStyle([("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#667eea")), ("TEXTCOLOR", (0, 0), (-1, 0), colors.white), ("GRID", (0, 0), (-1, -1), 0.25, colors.lightgrey), ("FONTSIZE", (0, 0), (-1, -1), 8)])))
    story.append(Spacer(1, 10))
    story.append(Paragraph(_escape(translate(locale, "reports.callsignActivity")), styles["Heading2"]))
    callsign_chart = _pdf_bar_chart(
        _report_callsign_chart_rows(report["callsigns"]),
        "report_label",
        "qso_count",
        170 * mm,
    )
    if callsign_chart is not None:
        story.append(callsign_chart)
        story.append(Spacer(1, 6))
    callsign_table = [[translate(locale, "reports.callsign"), translate(locale, "user.name"), translate(locale, "home.country"), translate(locale, "reports.qsos"), translate(locale, "reports.talkTime"), translate(locale, "reports.talkgroups")]] + [
        [str(row["callsign"]), str(row.get("source_name") or "—"), str(row.get("countries") or "—"), str(row["qso_count"]), _report_duration(row["duration_seconds"]), str(row["unique_talkgroups"])]
        for row in report["callsigns"]
    ]
    if len(callsign_table) == 1:
        callsign_table.append([translate(locale, "reports.noData"), "", "", "", "", ""])
    story.append(Table(callsign_table, repeatRows=1, style=TableStyle([("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#667eea")), ("TEXTCOLOR", (0, 0), (-1, 0), colors.white), ("GRID", (0, 0), (-1, -1), 0.25, colors.lightgrey), ("FONTSIZE", (0, 0), (-1, -1), 8)])))
    story.append(Spacer(1, 10))
    story.append(Paragraph(_escape(translate(locale, "reports.dailyTable")), styles["Heading2"]))
    daily_table = [[translate(locale, "reports.date"), translate(locale, "reports.qsos"), translate(locale, "reports.talkTime")]] + [[str(row["day"]), str(row["qso_count"]), _report_duration(row["duration_seconds"])] for row in report["daily"]]
    if len(daily_table) == 1:
        daily_table.append([translate(locale, "reports.noData"), "", ""])
    story.append(Table(daily_table, repeatRows=1, style=TableStyle([("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#667eea")), ("TEXTCOLOR", (0, 0), (-1, 0), colors.white), ("GRID", (0, 0), (-1, -1), 0.25, colors.lightgrey), ("FONTSIZE", (0, 0), (-1, -1), 8)])))
    document.build(story)
    return output.getvalue()


@router.get("/user/reports", response_class=HTMLResponse)
def user_reports(
    request: Request,
    timeRange: str = "1M",
    continent: str | None = None,
    country: str | None = None,
    talkgroup: list[int] | None = Query(default=None),
    callsign: str | None = None,
) -> Response:
    user = _current_user(request)
    if user is None:
        return RedirectResponse("/user/login", status_code=303)
    return _report_page(request, user, _report_time_range(timeRange), continent, country, talkgroup, callsign)


def _report_export(
    request: Request,
    time_range: str,
    continent: str | None,
    country: str | None,
    talkgroups: list[int] | None,
    callsign: str | None,
    kind: str,
) -> Response:
    user, report = _load_user_report(request, _report_time_range(time_range), continent, country, talkgroups, callsign)
    if user is None or report is None:
        return JSONResponse({"error": "authentication required"}, status_code=401)
    if kind == "csv":
        return Response(_report_csv(report), media_type="text/csv; charset=utf-8", headers={"Content-Disposition": f'attachment; filename="{user["callsign"]}-report.csv"'})
    if kind == "xlsx":
        return Response(_report_excel(report), media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers={"Content-Disposition": f'attachment; filename="{user["callsign"]}-report.xlsx"'})
    scope = callsign.strip() if callsign and callsign.strip() else translate(request_locale(request), "reports.allCallsigns")
    return Response(_report_pdf(report, scope, request_locale(request), histogram_bucket_seconds(_report_time_range(time_range))), media_type="application/pdf", headers={"Content-Disposition": f'attachment; filename="{user["callsign"]}-report.pdf"'})


@router.get("/user/reports/export.csv")
def user_report_csv(request: Request, timeRange: str = "1M", continent: str | None = None, country: str | None = None, talkgroup: list[int] | None = Query(default=None), callsign: str | None = None) -> Response:
    return _report_export(request, timeRange, continent, country, talkgroup, callsign, "csv")


@router.get("/user/reports/export.xlsx")
def user_report_excel(request: Request, timeRange: str = "1M", continent: str | None = None, country: str | None = None, talkgroup: list[int] | None = Query(default=None), callsign: str | None = None) -> Response:
    return _report_export(request, timeRange, continent, country, talkgroup, callsign, "xlsx")


@router.get("/user/reports/export.pdf")
def user_report_pdf(request: Request, timeRange: str = "1M", continent: str | None = None, country: str | None = None, talkgroup: list[int] | None = Query(default=None), callsign: str | None = None) -> Response:
    return _report_export(request, timeRange, continent, country, talkgroup, callsign, "pdf")



